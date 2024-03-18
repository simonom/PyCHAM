########################################################################
#								       #
# Copyright (C) 2018-2024					       #
# Simon O'Meara : simon.omeara@manchester.ac.uk			       #
#								       #
# All Rights Reserved.                                                 #
# This file is part of PyCHAM                                          #
#                                                                      #
# PyCHAM is free software: you can redistribute it and/or modify it    #
# under the terms of the GNU General Public License as published by    #
# the Free Software Foundation, either version 3 of the License, or    #
# (at  your option) any later version.                                 #
#                                                                      #
# PyCHAM is distributed in the hope that it will be useful, but        #
# WITHOUT ANY WARRANTY; without even the implied warranty of           #
# MERCHANTABILITY or## FITNESS FOR A PARTICULAR PURPOSE.  See the GNU  #
# General Public License for more details.                             #
#                                                                      #
# You should have received a copy of the GNU General Public License    #
# along with PyCHAM.  If not, see <http://www.gnu.org/licenses/>.      #
#                                                                      #
########################################################################
'''plots results from model and observations'''
# simulation and observation results are represented graphically

import matplotlib.pyplot as plt
from matplotlib.colors import BoundaryNorm
from matplotlib.ticker import MaxNLocator
# for customised colormap
from matplotlib.colors import LinearSegmentedColormap
# set colormap tick labels to standard notation
import matplotlib.ticker as ticker
import matplotlib.cm as cm
import os
import numpy as np
import scipy.constants as si
import openbabel.pybel as pybel
import openpyxl # for opening excel file

def plotter_gp_mod_n_obs(self): # for gas-phase concentration temporal profiles
	
	# open xls file
	wb = openpyxl.load_workbook(filename = self.xls_path)
	obs = wb['PyCHAM'] # get first sheet
	# empty list to hold information on handling observation data
	obs_setup = []
	# loop through columns to get information required for 
	# handling observational data
	for co in obs.iter_cols(values_only=True):
		obs_setup.append(co[0])
	
	try:
		# open worksheet with observations
		obs = wb[str(obs_setup[0])]
	except:
		mess = str('Error, could not find worksheet ' + str(obs_setup[0]) + ', available worksheets: ' + str(wb.sheetnames))
		self.l203a.setText(mess)
		return()

	# starting row for x-axis
	xrs = int(obs_setup[1].split(':')[0])-1
	# finishing row for x-axis
	xre = int(obs_setup[1].split(':')[1])-1
	# x-axis column
	xcol = int(obs_setup[2].split(':')[1])-1
	# starting row for y-axis
	yrs = int(obs_setup[6].split(':')[0])-1
	# finishing row for y-axis
	yre = int(obs_setup[6].split(':')[1])-1
	# starting column for y-axis
	ycs = int(obs_setup[7].split(':')[0])-1
	# finishing column for y-axis
	yce = int(obs_setup[7].split(':')[1])-1
	# empty array for x-axis data
	obsx = np.zeros((xre-xrs))
	# empty array for y-axis data
	obsy = np.zeros((yre-yrs, yce-ycs+1)) 
	cn = 0 # index counter
	xcn = 0 # x-axis counter
	ycn = 0 # y-axis counter
	# get x-axis data and y-axis data
	for co in obs.iter_rows(values_only=True):
		if (cn > xrs):
			# ensure compatability of types
			obsx = obsx.astype(type(co[xcol]))
			obsx[xcn] = co[xcol]
			obsy[ycn, :] = co[ycs:yce+1]
			xcn += 1 # x-axis counter
			ycn += 1 # y-axis counter
		cn += 1 # index counter
		if (cn > xre):
			break # stop loop through rows
	
	plt.ion() # show results to screen and turn on interactive mode
		
	# prepare plot
	fig, (ax0) = plt.subplots(1, 1, figsize=(14, 7))
	
	# legend labels of observed components
	llab = [lli for lli in obs_setup[9].split(',')]

	# loop through observed components to plot in order to plot
	for i in range(yce-ycs+1):
		
		if (self.gp_units[-4::] == 'near'): # linear y-axis
			eby = obsy[:, i]*0.0 # error bar array
			markers, caps, bars = ax0.errorbar(obsx, 
			obsy[:, i], yerr = eby, label = llab[i])
			# loop through error bars to set transparency
			[bar.set_alpha(0.1) for bar in bars]
		if (self.gp_units[-4::] == 'log.'): # logarithmic y-axis
			ax0.semilogy(obsx, obsy[:, i], label = llab[i])
	
	# ----------------------------------------------------------
	# deal with model results
	
	# get required variables from self
	wall_on = self.ro_obj.wf
	yrec = self.ro_obj.yrec
	num_comp = self.ro_obj.nc
	num_sb = self.ro_obj.nsb
	Nwet = self.ro_obj.Nrec_wet
	timehr = self.ro_obj.thr
	comp_names = self.ro_obj.names_of_comp
	rel_SMILES = self.ro_obj.rSMILES
	y_MW = (np.array((self.ro_obj.comp_MW))).reshape(1, -1)
	H2Oi = self.ro_obj.H2O_ind
	seedi = self.ro_obj.seed_ind
	indx_plot = self.ro_obj.plot_indx
	comp0 = self.ro_obj.init_comp
	rbou_rec = self.ro_obj.rad
	space_mode = self.ro_obj.spacing
	group_indx = self.ro_obj.gi

	# subtract any time before lights on
	
	timehr += float(obs_setup[4])

	# loop through modelled components
	for mci in range(len(self.gp_names)):
		if (self.gp_names[mci].strip() == ''): # if empty
			continue
		try:
			indx_plt = comp_names.index(self.gp_names[mci].strip())
		except:
			self.l203a.setText(str('Component ' + self.gp_names[mci].strip() + ' not found in chemical scheme used for this simulation'))
			# set border around error message
			if (self.bd_pl == 1):
				self.l203a.setStyleSheet(0., '2px dashed red', 0., 0.)
				self.bd_pl = 2
			else:
				self.l203a.setStyleSheet(0., '2px solid red', 0., 0.)
				self.bd_pl = 1

			plt.ioff() # turn off interactive mode
			plt.close() # close figure window
			return()
			
		if (self.gp_units[-4::] == 'near'): # linear y-axis
			ax0.plot(timehr, yrec[:, indx_plt], '--', label = str(self.gp_names[mci] + ' sim.'))
		if (self.gp_units[-4::] == 'log.'): # logarithmic y-axis
			ax0.semilogy(timehr, yrec[:, indx_plt], '--', label = str(self.gp_names[mci] + ' sim.'))
	
	# x-axis title
	ax0.set_xlabel(obs_setup[3], fontsize = 14)
	# y-axis title
	ax0.set_ylabel(obs_setup[8], fontsize = 14)
	# set label font size
	ax0.yaxis.set_tick_params(labelsize = 14, direction = 'in')
	ax0.xaxis.set_tick_params(labelsize = 14, direction = 'in')
	# include legend
	ax0.legend(fontsize = 14)

	# -----------------------------------------------------------

	return()

def plotter_pp_mod_n_obs(self): # for particle-phase concentration temporal profiles
	
	# open xls file
	wb = openpyxl.load_workbook(filename = self.xls_path)
	obs = wb['PyCHAM'] # get first sheet
	# empty list to hold information on handling observation data
	obs_setup = []
	# loop through columns to get information required for 
	# handling observational data
	for co in obs.iter_cols(values_only=True):
		obs_setup.append(co[0])
	
	# open worksheet with observations
	obs = wb[obs_setup[0]]
	xrs = int(obs_setup[1].split(':')[0])-1 # starting row for x-axis
	xre = int(obs_setup[1].split(':')[1])-1 # finishing row for x-axis
	xcol = int(obs_setup[2].split(':')[1])-1 # x-axis column
	yrs = int(obs_setup[6].split(':')[0])-1 # starting row for y-axis
	yre = int(obs_setup[6].split(':')[1])-1 # finishing row for y-axis
	ycs = int(obs_setup[7].split(':')[0])-1 # starting column for y-axis
	yce = int(obs_setup[7].split(':')[1])-1 # finishing column for y-axis
	obsx = np.zeros((xre-xrs)) # empty array for x-axis data
	obsy = np.zeros((yre-yrs, yce-ycs+1)) # empty array for y-axis data
	cn = 0 # index counter
	xcn = 0 # x-axis counter
	ycn = 0 # y-axis counter
	# get x-axis data and y-axis data
	for co in obs.iter_rows(values_only=True):
		if cn > xrs:
			obsx[xcn] = co[xcol]
			obsy[ycn, :] = co[ycs:yce+1]
			xcn += 1 # x-axis counter
			ycn += 1 # y-axis counter
		cn += 1 # index counter
		if cn> xre:
			break # stop loop through rows
		
	# legend labels of observed components
	llab = [lli for lli in obs_setup[9].split(',')]
	
	# ----------------------------------------------------------
	# deal with model results
	
	# get required variables from self
	wall_on = self.ro_obj.wf
	yrec = np.zeros((self.ro_obj.yrec.shape[0], self.ro_obj.yrec.shape[1]))
	yrec[:, :] = self.ro_obj.yrec[:, :]
	num_comp = self.ro_obj.nc
	num_sb = self.ro_obj.nsb
	Nwet = np.zeros((self.ro_obj.Nrec_wet.shape[0], self.ro_obj.Nrec_wet.shape[1]))
	Nwet[:, :] = self.ro_obj.Nrec_wet[:, :]
	Ndry = np.zeros((self.ro_obj.Nrec_dry.shape[0], self.ro_obj.Nrec_dry.shape[1]))
	Ndry[:, :] = self.ro_obj.Nrec_dry[:, :]
	timehr = self.ro_obj.thr
	comp_names = self.ro_obj.names_of_comp
	rel_SMILES = self.ro_obj.rSMILES
	y_MW = (np.array((self.ro_obj.comp_MW))).reshape(1, -1)
	H2Oi = self.ro_obj.H2O_ind
	seedi = self.ro_obj.seed_ind
	indx_plot = self.ro_obj.plot_indx
	comp0 = self.ro_obj.init_comp
	rbou_rec = np.zeros((self.ro_obj.rad.shape[0], self.ro_obj.rad.shape[1]))
	rbou_rec[:, :] = self.ro_obj.rad[:, :]
	space_mode = self.ro_obj.spacing

	# number of actual particle size bins
	num_asb = (num_sb-wall_on)

	
	plt.ion() # show results to screen and turn on interactive mode
		
	# prepare plot
	fig, (ax0) = plt.subplots(1, 1, figsize=(14, 7))

	# plotting against standard plot
	if (self.oandm == 1.1):	

		# prepare sub-plots depending on whether particles present
		if (num_asb > 0): 		

			# parasite axis setup on particle-phase plot-----------------------------
			par1 = ax0.twinx() # first parasite axis
			par2 = ax0.twinx() # second parasite axis
			# Offset the right spine of par2.  The ticks and label have already been
			# placed on the right by twinx above.
			par2.spines["right"].set_position(("axes", 1.2))
			# Having been created by twinx, par2 has its frame off, so the line of its
			# detached spine is invisible.  First, activate the frame but make the patch
			# and spines invisible.
			make_patch_spines_invisible(par2)
			# second, show the right spine
			par2.spines["right"].set_visible(True)	
			# ----------------------------------------------------------------------------------------

		# particle properties sub-plot --------------------------------------------------
		if (num_asb > 0): # if size bins present		

			if (timehr.ndim == 0): # occurs if only one time step saved
				Nwet = np.array(Nwet.reshape(1, num_asb))
				rbou_rec = np.array(rbou_rec.reshape(1, num_sb))
			if (num_asb == 1): # just one particle size bin (wall included in num_sb)
				Nwet = np.array(Nwet.reshape(len(timehr), num_asb))

			# plotting number size distribution --------------------------------------
		
			# don't use the first boundary as it could be zero, which will error when log10 taken
			log10D = np.log10(rbou_rec[:, 1::]*2.)
			
			if (num_asb > 1) :
				# note, can't append zero to start of log10D to cover first size bin as the log10 of the
				# non-zero boundaries give negative results due to the value being below 1, so instead
				# assume same log10 distance as the next pair
				log10D = np.append((log10D[:, 0]-(log10D[:, 1]-log10D[:, 0])).reshape(-1, 1), log10D, axis=1)
				# radius distance covered by each size bin (log10(um))
				dlog10D = (log10D[:, 1::]-log10D[:, 0:-1]).reshape(log10D.shape[0], log10D.shape[1]-1)
			
			if (num_asb == 1): # single particle size bin
				# assume lower radius bound is ten times smaller than upper
				dlog10D = (log10D[:, 0]-np.log10((rbou_rec[:, 1]/10.)*2.)).reshape(log10D.shape[0], 1)
					
			# number size distribution contours (# particles/cm3 (air))
			dNdlog10D = np.zeros((Nwet.shape[0], Nwet.shape[1]))
			dNdlog10D[:, :] = Nwet[:, :]/dlog10D[:, :]
			# transpose ready for contour plot
			dNdlog10D = np.transpose(dNdlog10D)
			
			# mask any nan values so they are not plotted
			z = np.ma.masked_where(np.isnan(dNdlog10D), dNdlog10D)
		
			# customised colormap (https://www.rapidtables.com/web/color/RGB_Color.html)
			colors = [(0.6, 0., 0.7), (0, 0, 1), (0, 1., 1.), (0, 1., 0.), (1., 1., 0.), (1., 0., 0.)]  # R -> G -> B
			n_bin = 100  # discretizes the colormap interpolation into bins
			cmap_name = 'my_list'
			# create the colormap
			cm = LinearSegmentedColormap.from_list(cmap_name, colors, N=n_bin)
		
			
			# -----------------------------
			# smallest value to consider for levels
			z_min = np.max(z[~np.isnan(z)])*1.e-3
		
			if np.max(z[~np.isnan(z)]) == 0: # if no particle present above 0 /cm3
				levels = np.zeros((1))
				
			else:	
				# make a first contour plot (which will be covered by plot p1 below) to get 
				# the contour labels of actual concentration (not log10(concentration))
				# set contour levels
				if (z_min > 0.5): # if rounding gives a number greater than 0 
					levels = np.arange(np.log10(round(z_min)), np.log10(np.max(z[~np.isnan(z)])), (np.log10(np.max(z[~np.isnan(z)]))-np.log10(round(z_min)))/1.e2)
				else: # i rounding would give a number below zero, then don't round		
					levels = np.arange(np.log10((z_min)), np.log10(np.max(z[~np.isnan(z)])), (np.log10(np.max(z[~np.isnan(z)]))-np.log10((z_min)))/1.e2)
				# associate colours and contour levels
				norm1 = BoundaryNorm(10.**levels, ncolors=cm.N, clip=True)
			
				# contour plot with times (hours) along x axis and 
				# particle diameters (nm) along y axis
				for ti in range(len(timehr)-1): # loop through times
					p0 = ax0.pcolormesh(timehr[ti:ti+2], (rbou_rec[ti, :]*2*1e3), (z[:, ti]).reshape(-1, 1), cmap=cm, norm=norm1)
			
			# ----------------------------
			
			if np.max(z[~np.isnan(z)]) == 0: # if no particle present above 0 /cm3
				
				levels = np.arange(-0.1, 0.1, (0.1--0.1)/1.e2)
				# associate colours and contour levels
				norm1 = BoundaryNorm(levels, ncolors=cm.N, clip=True)
				
				# contour plot with times (hours) along x axis and 
				# particle diameters (nm) along y axis
				for ti in range(len(timehr)-1): # loop through times
					p1 = ax0.pcolormesh(timehr[ti:ti+2], (rbou_rec[ti, :]*2*1e3), (z[:, ti]).reshape(-1, 1), cmap=cm, norm=norm1)

				cb = plt.colorbar(p1, format=ticker.FuncFormatter(fmt), pad=0.25, ax=ax0)
				
			else:
				if (z_min > 0.5): # if rounding would give zero
					# set contour levels
					levels = np.arange(np.log10(round(z_min)), np.log10(np.max(z[~np.isnan(z)])), (np.log10(np.max(z[~np.isnan(z)]))-np.log10(round(z_min)))/1.e2)
				else: # don't round if minimum close to zero
					levels = np.arange(np.log10((z_min)), np.log10(np.max(z[~np.isnan(z)])), (np.log10(np.max(z[~np.isnan(z)]))-np.log10((z_min)))/1.e2)
				
				# associate colours and contour levels
				norm1 = BoundaryNorm(levels, ncolors=cm.N, clip=True)
			
				# get indices of zeros in z
				zero_indx = z == 0.
				# get minimum value above 0 in z
				z_gt_zero = np.min(z[z!=0])
				# temporarily assign a > 0 number to zeros
				z[zero_indx] = z_gt_zero*1.e-1
				z_log10 = np.log10(z)
				z[zero_indx] = 0. # return to zero
			
				# contour plot with times (hours) along x axis and 
				# particle diameters (nm) along y axis
				for ti in range(len(timehr)-1): # loop through times
					p1 = ax0.pcolormesh(timehr[ti:ti+2], (rbou_rec[ti, :]*2*1e3), (z_log10[:, ti]).reshape(-1, 1), cmap=cm, norm=norm1)
		
				cb = plt.colorbar(p0, format=ticker.FuncFormatter(fmt), pad=0.25, ax=ax0)
		
			# if logarithmic or manual spacing of size bins specified, plot vertical axis 
			# logarithmically
			if space_mode == 'log' or space_mode == 'man':
				ax0.set_yscale("log")
				
			# set tick format for vertical axis
			ax0.yaxis.set_major_formatter(ticker.FormatStrFormatter('%.1e'))
			ax0.set_ylabel('Diameter (nm)', size = 14)
			ax0.xaxis.set_tick_params(labelsize = 14, direction = 'in', which = 'both')
			ax0.yaxis.set_tick_params(labelsize = 14, direction = 'in', which = 'both')

			# label according to whether gas-phase plot also displayed		
			if (indx_plot):
				ax0.text(x=timehr[0]-(timehr[-1]-timehr[0])/11., y = np.amax(rbou_rec*2*1e3)*1.05, s='b)', size=14)
			ax0.set_xlabel(r'Time through simulation (hours)', fontsize=14)
			
			cb.ax.tick_params(labelsize=14)   
			# colour bar label
			cb.set_label('dN (#$\,$$\mathrm{cm^{-3}}$)/d$\,$log$_{10}$(D$\mathrm{_p}$ ($\mathrm{\mu m}$))', size=14, rotation=270, labelpad=20)

			# ----------------------------------------------------------------------------------------
			# total particle number concentration # particles/cm3
		
			# include total number concentration (# particles/cm3 (air)) on contour plot
			# first identify size bins with radius exceeding 3nm
			# empty array for holding total number of particles
			Nvs_time = np.zeros((Nwet.shape[0]))
		
			for i in range(num_asb): # size bin loop
				Nvs_time[:] += Nwet[:, i] # sum number
			
			p3, = par1.plot(timehr, Nvs_time, '+k', label = 'N')
		
			par1.set_ylabel('N (#$\,$ $\mathrm{cm^{-3})}$', size=14, rotation=270, labelpad=20) # vertical axis label
			par1.yaxis.set_major_formatter(ticker.FormatStrFormatter('%.1e')) # set tick format for vertical axis
			par1.yaxis.set_tick_params(labelsize=14)

			# mass concentration of particles ---------------------------------------------------------------
			# array for mass concentration with time
			MCvst = np.zeros((1, len(timehr)))
			
			# first obtain just the particle-phase concentrations (# molecules/cm3)
			yrp = np.zeros((yrec.shape[0], num_comp*(num_asb)))
			yrp[:, :] = yrec[:, num_comp:num_comp*(num_asb+1)]
			# loop through size bins to convert to ug/m3
			for sbi in range(num_asb):
				yrp[:, sbi*num_comp:(sbi+1)*num_comp] = ((yrp[:, sbi*num_comp:(sbi+1)*num_comp]/si.N_A)*y_MW)*1.e12
			
			MCvst[0, :] = yrp.sum(axis=1)

			# log10 of maximum in mass concentration
			if (max(MCvst[0, :]) > 0):
				MCmax = int(np.log10(max(MCvst[0, :])))
			else:
				MCmax = 0.
		
			p5, = par2.plot(timehr, MCvst[0, :], 'xk', label = 'total sim. particle mass concentration')	
			
			# loop through observed components to plot in order to plot
			p6, = par2.plot(obsx-24., obsy[:, 0], '^k', label = llab[0])
			par2.set_ylabel(str('Mass Concentration ($\mathrm{\mu g\, m^{-3}})$'), rotation=270, size=16, labelpad=25)
			# set colour of label, tick font and corresponding vertical axis to match scatter plot presentation
			par2.yaxis.label.set_color('black')
			par2.tick_params(axis='y', colors='black')
			par2.spines['right'].set_color('black')
			par2.yaxis.set_major_formatter(ticker.FormatStrFormatter('%.1e')) # set tick format for vertical axis
			par2.yaxis.set_tick_params(labelsize=16)
			plt.legend(fontsize=14, handles=[p3, p5, p6] , loc=4, fancybox=True, framealpha=0.5)	

		# end of particle properties sub-plot -----------------------------------

		
		# x-axis title
		ax0.set_xlabel(obs_setup[3], fontsize = 14)
		
		# set label font size
		ax0.yaxis.set_tick_params(labelsize = 14, direction = 'in')
		ax0.xaxis.set_tick_params(labelsize = 14, direction = 'in')
		# include legend
		#ax0.legend(fontsize = 14)

		# -----------------------------------------------------------

		
	# plotting against cumulative particle mass concentration without water
	if (self.oandm == 1.2):	

		try: # in case user-supplied values given:
			psb_dub = [float(i) for i in self.e303p.text().split(',')]
		except:
			# default
			# particle size bins (upper bounds) diameter (um), 
			# ready for indexing the concentrations that will be summed
			psb_dub = [1.e-1, 5.e-1, 1.e0, 2.5e0, 1.e1, 2.e1]
		
		# convert diameters to nm from um
		for i in range(len(psb_dub)):
			psb_dub[i] = psb_dub[i]*1.e3

		# convert recorded radius to nm from um and from radius to diameter
		dbou_rec = self.ro_obj.rad*1.e3*2.
		
		# index array preparation
		psb_ind = np.zeros((len(self.ro_obj.thr), num_asb))

		# get the index of particle size bins
		for psb_dubi in psb_dub:
			
			psb_ind += dbou_rec[:, 1::]<psb_dubi
			
		# then use this index to find the mass concentrations 
		# inside each size bound

		# concentrations of components in particles (# molecules/cm3)
		yp = yrec[:, self.ro_obj.nc:self.ro_obj.nc*(num_asb+1)]
		# zero water
		yp[:, self.ro_obj.H2O_ind::self.ro_obj.nc] = 0

		# mass concentrations of each component in each size bin (ug/m3)
		mc = (yp/si.N_A)*np.tile(np.array((self.ro_obj.comp_MW)).reshape(1, -1),  (1, num_asb))*1e12

		# mass concentrations summed across components in each size bin (ug/m3)
		mc_psb = np.zeros((len(self.ro_obj.thr), num_asb))
		
		for i in range(num_asb): # loop through size bins
			mc_psb[:, i] = np.sum(mc[:, i*self.ro_obj.nc:(i+1)*self.ro_obj.nc], axis=1)

		# average mass per particle in this size bin (ug/m3/particle), note that components have been
		# allocated to size bins based on their wet (with water) sizes, and we want to display
		# the mass based on dry sizes
		mc_psb[self.ro_obj.Nrec_dry>0.] = mc_psb[self.ro_obj.Nrec_dry>0.]/self.ro_obj.Nrec_dry[self.ro_obj.Nrec_dry>0.]
		
		# now get mass integrated over all particles based on the dry (no water) number concentration
		# this will ensure the mass is allocated to the correct size bin
		mc_psb = mc_psb*self.ro_obj.Nrec_wet

		# results array preparation
		psb_res = np.zeros((len(self.ro_obj.thr), len(psb_dub)))

		# temporary holder array
		mc_psb_temp = np.zeros((len(self.ro_obj.thr), num_asb))

		for psb_dubi in range(len(psb_dub)): # loop through size categories
			
			# temporary holder array
			mc_psb_temp[:, :] = mc_psb[:, :]

			# zero the unwanted size bins
			mc_psb_temp[psb_ind<(len(psb_dub)-psb_dubi)] = 0.

			# get the concentrations of components in this size category (ug/m3)
			psb_res[:, psb_dubi] = np.sum(mc_psb_temp, axis=1)
		
			ax0.plot(self.ro_obj.thr, psb_res[:, psb_dubi], label = str(str(r'$D_{p}$<') + str(psb_dub[psb_dubi]/1.e3) + ' '+ str(r'$\rm{\mu}$m')))
			
		# now plot observations
		ax0.plot(obsx-24., obsy[:, 0], '^k', label = llab[0])

		ax0.set_ylabel(r'Concentration ($\rm{\mu}$g$\,$m$\rm{^{-3}}$)', fontsize = 14)
		ax0.set_xlabel(r'Time through simulation (hours)', fontsize = 14)
		ax0.yaxis.set_tick_params(labelsize = 14, direction = 'in')
		ax0.xaxis.set_tick_params(labelsize = 14, direction = 'in')
		
		# include legend
		ax0.legend(fontsize=14)

		

	return()


def plotter_VK_mod_n_obs(self): # for Van Krevelen diagrams
	
	# get required variables from self
	wall_on = self.ro_obj.wf
	yrec = self.ro_obj.yrec
	num_comp = self.ro_obj.nc
	num_sb = self.ro_obj.nsb
	Nwet = self.ro_obj.Nrec_wet
	timehr = self.ro_obj.thr
	comp_names = self.ro_obj.names_of_comp
	rel_SMILES = self.ro_obj.rSMILES
	y_MW = (np.array((self.ro_obj.comp_MW))).reshape(1, -1)
	H2Oi = self.ro_obj.H2O_ind
	seedi = self.ro_obj.seed_ind
	indx_plot = self.ro_obj.plot_indx
	comp0 = self.ro_obj.init_comp
	rbou_rec = self.ro_obj.rad
	space_mode = self.ro_obj.spacing
	
	self.HC = np.array((ro_obj.HyC)) # hydrogen:carbon ratios of each component


	if (self.oandm == 2 or self.oandm == 4 or self.oandm == 6): # if user wants average over Hydrocarbons
		# index of the required experiment period
		tindx = (timehr*3600. >= 0)*(timehr*3600. <= timehr[-1]*3600.)
		# isolate gas-phase concentrations during this period (ppb)
		yrel = yrec[tindx, 0:num_comp]
		
		if (self.oandm == 2): # all hydrocarbons wanted
			# index for all hydrocarbons
			HCindx = (self.HC > 0.)
		
		# non-methane hydrocarbons wanted
		if (self.oandm == 4): # non-methane hydrocarbons wanted
			
			Ccount = [] # count number of carbons in each component
			for comp in rel_SMILES:
				Ccount.append(comp.count('C')+comp.count('c'))

			# index for non-methane hydrocarbons
			HCindx = (self.HC > 0.)*(np.array((Ccount)) > 1)

		# _ extension hydrocarbons wanted
		if (self.oandm == 6): # _ extension hydrocarbons wanted
			
			ext_cnt = [] # count number of carbons in each component
			for comp in comp_names:
				if comp.count('_') > 0:
					ext_cnt.append(1)
				else:
					ext_cnt.append(0)

			# index for _ extension hydrocarbons
			HCindx = (np.array((ext_cnt)) > 0)

		# isolate hydrocarbon molecules
		yrel = yrel[:, HCindx]
		# sum of concentrations
		HCsum = yrel.sum(axis=1).reshape(-1, 1)

		# fractional gas-phase concentration
		z_indx = (HCsum > 0.).reshape(-1, 1)

		yrel_frac = yrel[z_indx[:, 0], :]/((HCsum[z_indx]).reshape(-1, 1))
		OC_rel = (np.array((OC)))[HCindx].reshape(1, -1) # and ensure correct alignment with yrel
		HC_rel = self.HC[HCindx].reshape(1, -1) # and ensure correct alignment with yrel
		# weight concentrations by O:C ratio
		OCweight = yrel_frac*OC_rel
		# weight concentrations by H:C ratio
		HCweight = yrel_frac*HC_rel
	
		# get the average gas-phase O:C and H:C ratios during the stated period
		OCav = OCweight.sum(axis=1)
		HCav = HCweight.sum(axis=1)

		plt.ion() # show results to screen and turn on interactive mode
		
		# prepare plot
		fig, (ax0) = plt.subplots(1, 1, figsize=(14, 7))

		# colormap from library (https://matplotlib.org/stable/api/cm_api.html#matplotlib.cm.get_cmap)
		colors = cm.get_cmap('ocean')

		# set contour levels
		levels = (MaxNLocator(nbins = 256).tick_values(np.min(timehr[tindx]*3600), np.max(timehr[tindx]*3600)))
	
		# associate colours and contour levels
		norm1 = BoundaryNorm(levels, ncolors=256, clip=True)
	
		for i in range(len(OCav)):
			ax0.plot(OCav[i], HCav[i], 'o', mec = 'k', mfc = colors(i/(len(OCav)-1)))

		# main title
		if (self.oandm == 2):
			ax0.set_title('All Hydrocarbons')
		if (self.oandm == 4):
			ax0.set_title('Non-methane Hydrocarbons')
		if (self.oandm == 6):
			ax0.set_title('Hydrocarbons Containing _ in Name')
		# x-axis title
		ax0.set_xlabel('O:C ratio number-averaged over hydrocarbons', fontsize = 14)
		# y-axis title
		ax0.set_ylabel('H:C ratio number-averaged over hydrocarbons', fontsize = 14)
		# set label font size
		ax0.yaxis.set_tick_params(labelsize = 14, direction = 'in')
		ax0.xaxis.set_tick_params(labelsize = 14, direction = 'in')

		# include colorbar to demonstrate time through experiment (s)
		cb = fig.colorbar(cm.ScalarMappable(norm=norm1, cmap=colors), ax=ax0)

		cb.ax.tick_params(labelsize=14)   
		# colour bar label
		cb.set_label('Time since experiment start (s)', size=14, rotation=270, labelpad=20)

	if (self.oandm == 3 or self.oandm == 5 or self.oandm == 7): # if user wants individual hydrocarbons at a given time

		# get the time through experiment (s) in question
		try:
			treq = float((self.e403.toPlainText()))
		except:
			treq = timehr[-1]*3600.

		# index of the required experiment period
		tindx = np.min(np.abs(timehr*3600-treq)) == (np.abs(timehr*3600-treq))

		# isolate gas-phase concentrations during this period (ppb)
		yrel = yrec[tindx, 0:num_comp]

		if (self.oandm == 3): # all individual hydrocarbons
			HCindx = self.HC > 0.

		if (self.oandm == 5): # non-methane individual hydrocarbons
			Ccount = [] # count number of carbons in each component
			for comp in rel_SMILES:
				Ccount.append(comp.count('C')+comp.count('c'))

			# index for non-methane hydrocarbons
			HCindx = (self.HC > 0.)*(np.array((Ccount)) > 1)

				# _ extension hydrocarbons wanted

		if (self.oandm == 7): # _ extension hydrocarbons wanted
			
			ext_cnt = [] # count number of carbons in each component
			for comp in comp_names:
				if comp.count('_') > 0:
					ext_cnt.append(1)
				else:
					ext_cnt.append(0)

			# index for _ extension hydrocarbons
			HCindx = (np.array((ext_cnt)) > 0)

		# isolate hydrocarbon molecules
		yrel = yrel[0, HCindx]

		# isolate H:C and O:C of hydrocarbons
		OC = (np.array(OC))[HCindx]
		HC = (np.array(self.HC))[HCindx]
		comp_names = (np.array(comp_names))[HCindx]
		
		# sum of these hydrocarbons
		HC_sum = np.sum(yrel)

		if (HC_sum > 0.):
			# normalise (0-1) hydrocarbon mixing ratios
			yrel = yrel/HC_sum
		else:
			yrel = []
			OC = []
			HC = []

		plt.ion() # show results to screen and turn on interactive mode
		
		# prepare plot
		fig, (ax0) = plt.subplots(1, 1, figsize=(14, 7))

		# colormap from library (https://matplotlib.org/stable/api/cm_api.html#matplotlib.cm.get_cmap)
		colors = cm.get_cmap('ocean')

		# set contour levels
		levels = (MaxNLocator(nbins = 256).tick_values(0., 1.))
	
		# associate colours and contour levels
		norm1 = BoundaryNorm(levels, ncolors=256, clip=True)
	
		for i in range(len(HC)):
			ax0.plot(OC[i], HC[i], 'o', mec = 'k', mfc = colors(yrel[i]))
			
		# plot title
		if (self.oandm == 3):
			ax0.set_title(str('Van Krevelen at ' + str(treq) + ' s since experiment start for all hydrocarbons'), fontsize = 14)
		if (self.oandm == 5):
			ax0.set_title(str('Van Krevelen at ' + str(treq) + ' s since experiment start for non-methane hydrocarbons'), fontsize = 14)
		if (self.oandm == 7):
			ax0.set_title(str('Van Krevelen at ' + str(treq) + ' s since experiment start for extension hydrocarbons'), fontsize = 14)
		
		# x-axis title
		ax0.set_xlabel('O:C ratio of hydrocarbon', fontsize = 14)
		# y-axis title
		ax0.set_ylabel('H:C ratio of hydrocarbon', fontsize = 14)
		# set label font size
		ax0.yaxis.set_tick_params(labelsize = 14, direction = 'in')
		ax0.xaxis.set_tick_params(labelsize = 14, direction = 'in')

		# include colorbar to demonstrate time through experiment (s)
		cb = fig.colorbar(cm.ScalarMappable(norm=norm1, cmap=colors), ax=ax0)
		cb.ax.tick_params(labelsize=14)   
		# colour bar label
		cb.set_label('Normalised number concentration at this time', size=14, rotation=270, labelpad=20)


	return()

def plotter_mass_defect(self): # for mass defect plots

	# get required variables from self
	wall_on = self.ro_obj.wf
	yrec = self.ro_obj.yrec
	num_comp = self.ro_obj.nc
	num_sb = self.ro_obj.nsb
	Nwet = self.ro_obj.Nrec_wet
	timehr = self.ro_obj.thr
	comp_names = self.ro_obj.names_of_comp
	rel_SMILES = self.ro_obj.rSMILES
	y_MW = (np.array((self.ro_obj.comp_MW))).reshape(1, -1)
	H2Oi = self.ro_obj.H2O_ind
	seedi = self.ro_obj.seed_ind
	indx_plot = self.ro_obj.plot_indx
	comp0 = self.ro_obj.init_comp
	rbou_rec = self.ro_obj.rad
	space_mode = self.ro_obj.spacing
	
	# get exact molar masses (g/mol)
	y_MW = np.array((y_MW))

	# get nominal mass (g/mol)
	nom_mass = np.array((ro_obj.nominal_mass))

	# calculate the mass defect according to doi.org/10.1002/jms.2953226 (g/mol)
	mass_def = y_MW-nom_mass

	plt.ion() # show results to screen and turn on interactive mode
		
	# prepare plot
	fig, (ax0) = plt.subplots(1, 1, figsize=(14, 7))

	if (self.oandm == 8): # all components in chemical scheme without scaling

		# plot mass defect against nominal mass
		ax0.plot(nom_mass, mass_def, '+')

	# all components in chemical scheme scaled to concentration at provided time
	if (self.oandm == 9):

		self.HC = np.array((ro_obj.HyC)) # hydrogen:carbon ratios of each component

		# get the time through experiment (s) in question
		try:
			treq = float((self.e403.toPlainText()))
		except:
			treq = timehr[-1]*3600.

		# index of the required experiment period
		tindx = np.min(np.abs(timehr*3600-treq)) == (np.abs(timehr*3600-treq))

		# index of hydrocarbons
		HCindx = self.HC > 0.

		# isolate gas-phase concentrations of hydrocarbons at this time (ppb)
		yrel = (yrec[tindx, 0:num_comp])[0, HCindx]

		# normalise to sum of concentrations
		yrel = yrel/sum(yrel)

		# colormap from library (https://matplotlib.org/stable/api/cm_api.html#matplotlib.cm.get_cmap)
		colors = cm.get_cmap('ocean')

		# set contour levels
		levels = (MaxNLocator(nbins = 256).tick_values(np.min(yrel), np.max(yrel)))
	
		# associate colours and contour levels
		norm1 = BoundaryNorm(levels, ncolors=256, clip=True)
	
		for i in range(len(yrel)):
			ax0.plot(nom_mass[HCindx][i], mass_def[HCindx][i], 'o', mec = 'k', mfc = colors(yrel[i]))

		# include colorbar to demonstrate time through experiment (s)
		cb = fig.colorbar(cm.ScalarMappable(norm=norm1, cmap=colors), ax=ax0)
		cb.ax.tick_params(labelsize=14)   
		# colour bar label
		cb.set_label('Normalised number concentration at this time', size=14, rotation=270, labelpad=20)

		ax0.set_title(str('Mass Defect at ' + str(treq) + ' s since experiment start for all hydrocarbons'), fontsize = 14)
			
	# x-axis title
	ax0.set_xlabel('Nominal Molar Mass (g/mol)', fontsize = 14)
	# y-axis title
	ax0.set_ylabel('Mass Defect (Exact Mass-Nominal Mass) (g/mol)', fontsize = 14)
	# set label font size
	ax0.yaxis.set_tick_params(labelsize = 14, direction = 'in')
	ax0.xaxis.set_tick_params(labelsize = 14, direction = 'in')

	return()
	
def plotter_exp_prep(self): # for experiment design

	# get required variables from self
	wall_on = self.ro_obj.wf
	yrec = self.ro_obj.yrec
	num_comp = self.ro_obj.nc
	num_sb = self.ro_obj.nsb
	Nwet = self.ro_obj.Nrec_wet
	timehr = self.ro_obj.thr
	comp_names = self.ro_obj.names_of_comp
	rel_SMILES = self.ro_obj.rSMILES
	y_MW = (np.array((self.ro_obj.comp_MW))).reshape(1, -1)
	H2Oi = self.ro_obj.H2O_ind
	seedi = self.ro_obj.seed_ind
	indx_plot = self.ro_obj.plot_indx
	comp0 = self.ro_obj.init_comp
	rbou_rec = self.ro_obj.rad
	space_mode = self.ro_obj.spacing
	Cfac = self.ro_obj.cfac
	group_indx = self.ro_obj.gi
	
	# simulation title ---------------------------------------------
	sim_title = self.dir_path
	# -----------------------------------------------------------------
	
	# -----------------------------------------------------------------
	rationale = [] # remember to leave column for rationale
	# -----------------------------------------------------------------
	
	# -----------------------------------------------------------------
	# ratio of HO2, NO and RO2 reactivity with PINALO2 
	# averaged over whole simulation
	
	# prepare to hold values
	ct_PINALO2 = np.zeros((3))
	
	# name of file that would contain change tendencies of PINALO2
	fname = str(self.dir_path+ '/PINALO2_rate_of_change')
	
	try: # try to open
		dydt = np.loadtxt(fname, delimiter = ',', skiprows = 0) # skiprows = 0 includes header
		# isolate chemical reaction numbers, note last two columns for gas-particle and gas-wall partitioning
		dydt_header = dydt[0, 0:-2]	
		dydt = dydt[1::, :] # exclude header now
		
	except: # if unable to open file
		PINALO2_ct = str('No change tendency record for PINALO2 was found, was it specified in the tracked_comp input of the model variables file?  Please see README for more information.')
	
	# prepare to store results of change tendency due to chemical reactions
	res = np.zeros((dydt.shape[0], dydt.shape[1]-2))
	res[:, :] = dydt[:, 0:-2] # get chemical reaction numbers and change tendencies
		
	import scipy.constants as si
	
	# get all reactions out of the used chemical scheme --------------------------	
	import sch_interr # for interpeting chemical scheme
	import re # for parsing chemical scheme
	import scipy.constants as si

	sch_name = self.ro_obj.sp
	inname = self.ro_obj.vp
	
	f_open_eqn = open(sch_name, mode= 'r' ) # open model variables file
	# read the file and store everything into a list
	total_list_eqn = f_open_eqn.readlines()
	f_open_eqn.close() # close file
		
	inputs = open(inname, mode= 'r' ) # open model variables file
	in_list = inputs.readlines() # read file and store everything into a list
	inputs.close() # close file
	
	# default chemical scheme markers
	self.chem_sch_mrk = ['{', 'RO2', '+', 'C(ind_', ')','' , '&', '' , '', ':', '}', ';']
	
	for i in range(len(in_list)): # loop through supplied model variables to interpret

		# ----------------------------------------------------
		# if commented out continue to next line
		if (in_list[i][0] == '#'):
			continue
		key, value = in_list[i].split('=') # split values from keys
		# model variable name - a string with bounding white space removed
		key = key.strip()
		# ----------------------------------------------------

		if key == 'chem_scheme_markers' and (value.strip()): # formatting for chemical scheme
			self.chem_sch_mrk = [str(i).strip() for i in (value.split(','))]

	# interrogate scheme to list equations
	[eqn_list, aqeqn_list, eqn_num, rrc, rrc_name, 
		RO2_names] = sch_interr.sch_interr(total_list_eqn, self)	
	
	# list interested reactants
	reac_interest = [' HO2 ', '*RO2', ' NO ']
	
	eqn_cnt = 0 # count on equations through chemical scheme
	PINALO2_eqn_cnt = 0 # count on PINALO2 equations
	
	# loop through equations to find where PINALO2 reacts with HO2, RO2 or NO
	for eqn_numi in dydt_header:
		
		# obtain original equation
		eqni = eqn_list[int(eqn_numi)]
		
		for reaci in range(3): # loop through the reactants we're interested in
		
			# focus on LHS of reaction first
			if reac_interest[reaci] in eqni.split('=')[0]:
				# sum change tendency over time and add to total
				ct_PINALO2[reaci] += np.sum(dydt[:, PINALO2_eqn_cnt])
				break # move onto next equation
			
			# the reaction rate coefficient for this equation
			# where reaction rate coefficient starts
			indx_rrc_start = eqni.index(self.chem_sch_mrk[9])
			# where equation starts
			indx_rrc_end = eqni.index(self.chem_sch_mrk[10])
			
			# in case equation comes before the reaction rate coefficient,
			# then change end point of reaction rate coefficient part to end 
			# of line
			if (indx_rrc_end < indx_rrc_start):
				indx_rrc_end = eqni.index(self.chem_sch_mrk[11])
				
			# likewise, if reactant is in the reaction rate coefficient
			if reac_interest[reaci] in eqni[indx_rrc_start:indx_rrc_end]:
				
				# sum change tendency over time and add to total
				ct_PINALO2[reaci] += np.sum(dydt[:, PINALO2_eqn_cnt])
				break
		
		PINALO2_eqn_cnt+= 1 # count on equations

	# normalise change tendencies of PINALO2 to the minimum change tendency
	ct_PINALO2 = np.abs(ct_PINALO2[:])/min(np.abs(ct_PINALO2[ct_PINALO2!=0.]))

	# round to two significant figures
	for i in range(len(ct_PINALO2)):
		ct_PINALO2[i] = float('%s' % float('%.2g' % ct_PINALO2[i]))

	# -----------------------------------------------------------------

	# open an excel file ready for saving results
	from openpyxl import Workbook
	wb = Workbook() # create new workbook
	ws = wb.active # get active worksheet
	ws.title = "Exp. Prep." # set worksheet title
	# add headers
	ws["A1"] = 'Exp. #'
	ws["B1"] = 'Rationale'
	ws["N1"] = 'HO2:NO:RO2 reactivity with PINALO2 over whole simulation'
	# add data
	ws["A2"] = 1
	ws["B2"] = 'HO2 and RO2 at play'
	ws["N2"] = str(str(ct_PINALO2[0]) + ':' + str(ct_PINALO2[2]) + ':' + str(ct_PINALO2[1]))
	wb.save(filename = str(self.dir_path+ '/exp_prep.xlsx')) # save workbook

	# -----------------------------------------------------------------
	# ratio of O3, OH and NO3 reactivity with APINENE 
	# averaged over whole simulation
	
	# prepare to hold values
	ct_APINENE = np.zeros((3))
	
	# name of file that would contain change tendencies of PINALO2
	fname = str(self.dir_path+ '/APINENE_rate_of_change')
	
	try: # try to open
		dydt = np.loadtxt(fname, delimiter = ',', skiprows = 0) # skiprows = 0 includes header
		# isolate chemical reaction numbers, note last two columns for gas-particle and gas-wall partitioning
		dydt_header = dydt[0, 0:-2]	
		dydt = dydt[1::, :] # exclude header now
		
	except: # if unable to open file
		APINENE_ct = str('No change tendency record for APINENE was found, was it specified in the tracked_comp input of the model variables file?  Please see README for more information.')
	
	# prepare to store results of change tendency due to chemical reactions
	res = np.zeros((dydt.shape[0], dydt.shape[1]-2))
	res[:, :] = dydt[:, 0:-2] # get chemical reaction numbers and change tendencies
		
	import scipy.constants as si
	
	# get all reactions out of the used chemical scheme --------------------------	
	import sch_interr # for interpeting chemical scheme
	import re # for parsing chemical scheme
	import scipy.constants as si

	sch_name = self.ro_obj.sp
	inname = self.ro_obj.vp
	
	f_open_eqn = open(sch_name, mode= 'r' ) # open model variables file
	# read the file and store everything into a list
	total_list_eqn = f_open_eqn.readlines()
	f_open_eqn.close() # close file
		
	inputs = open(inname, mode= 'r' ) # open model variables file
	in_list = inputs.readlines() # read file and store everything into a list
	inputs.close() # close file
	
	# default chemical scheme markers
	self.chem_sch_mrk = ['{', 'RO2', '+', 'C(ind_', ')','' , '&', '' , '', ':', '}', ';']
	
	for i in range(len(in_list)): # loop through supplied model variables to interpret

		# ----------------------------------------------------
		# if commented out continue to next line
		if (in_list[i][0] == '#'):
			continue
		key, value = in_list[i].split('=') # split values from keys
		# model variable name - a string with bounding white space removed
		key = key.strip()
		# ----------------------------------------------------

		if key == 'chem_scheme_markers' and (value.strip()): # formatting for chemical scheme
			self.chem_sch_mrk = [str(i).strip() for i in (value.split(','))]

		# for later on in this function, look for, and save physical variables 
		# (RH, Temperature, Pressure, Light Status)
		if key == 'rh' and (value.strip()): # relative humidity in chamber (0-1)
				self.RH = (np.array(([float(i) for i in ((value.strip()).split(','))])))[0]
		
		if key == 'temperature' and (value.strip()): # chamber temperature (K)
				self.TEMP = [float(i) for i in ((value.strip()).split(','))][0]	

		if key == 'p_init' and (value.strip()): # pressure inside chamber
				self.Press = float(value.strip())

		if key == 'light_status' and value.strip(): # status of lights (on or off)
				light_stat = [int(i) for i in (value.split(','))]
				self.light_stat = np.max(np.array((light_stat)))

	# interrogate scheme to list equations
	[eqn_list, aqeqn_list, eqn_num, rrc, rrc_name, 
		RO2_names] = sch_interr.sch_interr(total_list_eqn, self)	
	
	# list interested reactants
	reac_interest = [' O3 ', ' OH ', ' NO3 ']
	
	eqn_cnt = 0 # count on equations through chemical scheme
	APINENE_eqn_cnt = 0 # count on APINENE equations
	
	# loop through equations to find where APINENE reacts with O3, OH or NO3
	for eqn_numi in dydt_header:
		
		# obtain original equation
		eqni = eqn_list[int(eqn_numi)]
		
		for reaci in range(len(reac_interest)): # loop through the reactants we're interested in
		
			# focus on LHS of reaction first
			if reac_interest[reaci] in eqni.split('=')[0]:
				# sum change tendency over time and add to total
				ct_APINENE[reaci] += np.sum(dydt[:, APINENE_eqn_cnt])
				break # move onto next equation
			
			# the reaction rate coefficient for this equation
			# where reaction rate coefficient starts
			indx_rrc_start = eqni.index(self.chem_sch_mrk[9])
			# where equation starts
			indx_rrc_end = eqni.index(self.chem_sch_mrk[10])
			
			# in case equation comes before the reaction rate coefficient,
			# then change end point of reaction rate coefficient part to end 
			# of line
			if (indx_rrc_end < indx_rrc_start):
				indx_rrc_end = eqni.index(self.chem_sch_mrk[11])
				
			# likewise, if reactant is in the reaction rate coefficient
			if reac_interest[reaci] in eqni[indx_rrc_start:indx_rrc_end]:
				
				# sum change tendency over time and add to total
				ct_APINENE[reaci] += np.sum(dydt[:, APINENE_eqn_cnt])
				break
		
		APINENE_eqn_cnt+= 1 # count on equations

	# normalise change tendencies of APINENE to the minimum change tendency
	ct_APINENE = np.abs(ct_APINENE[:])/min(np.abs(ct_APINENE[ct_APINENE!=0.]))

	# round to two significant figures
	for i in range(len(ct_APINENE)):
		ct_APINENE[i] = float('%s' % float('%.2g' % ct_APINENE[i]))

	# -----------------------------------------------------------------

	# add headers to excel file
	ws["O1"] = 'O3:OH:NO3 reactivity with APINENE over whole simulation'
	# add data
	ws["O2"] = str(str(ct_APINENE[0]) + ':' + str(ct_APINENE[1]) + ':' + str(ct_APINENE[2]))

	# -----------------------------------------------------------------
	# now onto initial concentrations (ppb) of alpha-pinene, O3, H2O2, NO, CO, CH4
	
	# list of components we want initial concentrations of
	init_comp = ['APINENE', 'O3', 'H2O2', 'NO', 'CO', 'CH4']
	
	# spreadsheet columns to use 
	col = ['C', 'D', 'E', 'F', 'G', 'H']

	for i in range(len(init_comp)): # loop through components

		# add header to excel file
		cell_id = str(col[i]+'1') # cell ID for header
		ws[cell_id] = str('[' + init_comp[i] + '] (ppb) at t=0')

		cell_id = str(col[i]+'2') # cell ID for value
		# get initial gas-phase concentration (ppb)
		conc_now = yrec[0, comp_names.index(init_comp[i])]
		# round to two significant figures
		conc_now = float('%s' % float('%.2g' % conc_now))

		conc_now = "{:e}".format(conc_now) # ensure scientific notation
		if len(conc_now)>conc_now.index('.')+1:
			conc_now = str(conc_now[0:conc_now.index('.')+2] + conc_now[conc_now.index('e')::])

		# store result
		ws[cell_id] = conc_now

	# now onto components whose concentration we want to know --------- 
	# averaged over the experiment
	
	# list of components we want average concentrations of
	aver_comp = ['HO2', 'OH', 'RO2', 'CH3O2']

	# spreadsheet columns to use 
	col = ['I', 'J', 'K', 'L']
	
	for i in range(len(aver_comp)): # loop through components


		# if the short-lived component, then present in units of # molecules/cm3
		if (aver_comp[i] == 'HO2' or aver_comp[i] == 'OH'):
			# add header to excel file
			cell_id = str(col[i]+'1') # cell ID for header
			ws[cell_id] = str('[' + aver_comp[i] + '] (# molecules/cm3) geometric mean over whole experiment')
			# get concentration and note conversion from ppb to # molecules/cm3
			conc_now = np.sum(yrec[:, comp_names.index(aver_comp[i])])/yrec.shape[0]*Cfac[0]


		else: # if longer-lived components
			if (aver_comp[i] == 'RO2'):
				indx_plot = (np.array((group_indx['RO2i'])))
				conc_now = yrec[:, indx_plot].sum(axis=1)
				# remember geometric mean average of RO2 for CH3O2:RO2 ratio 
				# later
				conc_RO2 = np.sum(yrec[:, indx_plot].sum(axis=1))/yrec.shape[0]
			else: # if individual components
				conc_now = yrec[:, comp_names.index(aver_comp[i])]
				# remember geometric mean average for CH3O2 for CH3O2:RO2 
				# ratio later
				if (aver_comp[i] == 'CH3O2'):
					conc_CH3O2 = np.sum(yrec[:, comp_names.index(aver_comp[i])])/yrec.shape[0]

			# add header to excel file
			cell_id = str(col[i]+'1') # cell ID for header
			ws[cell_id] = str('[' + aver_comp[i] + '] (ppb) geometric mean over whole experiment')
			# get geometric mean concentration (ppb)
			conc_now = np.sum(conc_now)/yrec.shape[0]

		cell_id = str(col[i]+'2') # cell ID for value
		
		# round to two significant figures
		conc_now = float('%s' % float('%.2g' % conc_now))

		conc_now = "{:e}".format(conc_now) # ensure scientific notation
		if len(conc_now)>conc_now.index('.')+1:
			conc_now = str(conc_now[0:conc_now.index('.')+2] + conc_now[conc_now.index('e')::])
		
		# store result
		ws[cell_id] = conc_now

	# ratio of concentration of CH3O2 to RO2 ------------------
	RO2_ratio = conc_RO2/conc_CH3O2 

	# round to two significant figures
	RO2_ratio = float('%s' % float('%.2g' % RO2_ratio))

	RO2_ratio = "{:e}".format(RO2_ratio) # ensure scientific notation
	if len(RO2_ratio)>RO2_ratio.index('.')+1:
		RO2_ratio = str(RO2_ratio[0:RO2_ratio.index('.')+2] + RO2_ratio[RO2_ratio.index('e')::])
	
	# header for result
	ws["M1"] = str('[CH3O2]:[RO2] geometric mean over whole simulation')

	# store result
	ws["M2"] = str('1:' + str(RO2_ratio))

	# include physical variables of experiment ----------------
	# RH, temperature, pressure, light
	
	# header for result
	ws["P1"] = str('RH (%)')
	ws["P2"] = str(self.RH*100.)
	
	ws["Q1"] = str('Temperature (K)')
	ws["Q2"] = str(self.TEMP)
		
	ws["R1"] = str('Pressure (Pa)')
	ws["R2"] = str(self.Press)
		
	ws["S1"] = str('Light Status')
	if (self.light_stat == 0):
		ws["S2"] = 'Off'
	if (self.light_stat == 1):
		ws["S2"] = 'On'

	wb.save(filename = str(self.dir_path+ '/exp_prep.xlsx')) # save workbook
	return()

# function to makes spines invisible
def make_patch_spines_invisible(ax):
	ax.set_frame_on(True)
	ax.patch.set_visible(False)
	for sp in ax.spines.values():
		sp.set_visible(False)

# function for doing colorbar tick labels in standard notation
def fmt(x, pos):
	a, b = '{:.1e}'.format(x).split('e')
	b = int(b)
	return r'${} \times 10^{{{}}}$'.format(a, b)
 
