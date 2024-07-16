########################################################################
#								       #
# Copyright (C) 2018-2024					       #
# Simon O'Meara : simon.omeara@manchester.ac.uk			       #
#								       #
# All Rights Reserved.                                                 #
# This file is part of PyCHAM                                          #
#                                                                      #
# PyCHAM is free software: you can redistribute it and/or modify it    #
# under the terms of the GNU General Public License as published by    #
# the Free Software Foundation, either version 3 of the License, or    #
# (at  your option) any later version.                                 #
#                                                                      #
# PyCHAM is distributed in the hope that it will be useful, but        #
# WITHOUT ANY WARRANTY; without even the implied warranty of           #
# MERCHANTABILITY or## FITNESS FOR A PARTICULAR PURPOSE.  See the GNU  #
# General Public License for more details.                             #
#                                                                      #
# You should have received a copy of the GNU General Public License    #
# along with PyCHAM.  If not, see <http://www.gnu.org/licenses/>.      #
#                                                                      #
########################################################################
'''opening and reading any user-provided observation file'''
# called from eqn_pars.py and (if in simulation preparation mode) 
# ui_check.py

# define function
def obs_file_open(self):

	# inputs: --
	# self - PyCHAM object
	# ----------

	import openpyxl
	import os
	import numpy as np

	self.obs_file = str(self.obs_file)
	try: # in case full path provided
		wb = openpyxl.load_workbook(filename = self.obs_file)
	except: # in case in same folder as model variables file
		pd_indx = self.inname[::-1].index('/')
		pd = self.inname[0:-pd_indx]
		self.obs_file = str(pd + self.obs_file)
		wb = openpyxl.load_workbook(filename = self.obs_file)
	sheet = wb['PyCHAMobs']
	# time (seconds) is in first column, other 
	# variables in later columns
	ic = 0 # count on row iteration
	# column indices for temperature and 
	# relative humidity
	temper_indx = -2
	rh_indx = -2
	# track whether particulate matter mole fractions 
	# seen in obs_file
	seedx_flag = 0
	# track whether PM number concentrations
	# seen in obs_file
	pconc_flag = 0
	# track whether mean radius seen
	mean_rad_flag = 0
	# size bin names
	sb_nam = []
	# names of particulate matter component
	pm_comp = []
	# column indices of PM mole fractions
	pm_col = []
	# column indices for PM number concentration
	pconc_col = []
	# column indices for PM mean radius
	mean_rad_col = []
	# size bin names for pconc
	sb_nam_pconc = []
	# size bin names for mean_rad
	sb_nam_mean_rad = []
	
	# index for columns of components
	comp_indx = []
	# loop through rows
	for i in sheet.iter_rows(values_only = True):
		if (ic == 0):
			# get chemical scheme names of variables
			names_xlsx = i[1::]
			# prepare for variable names
			self.obs_comp = []
			# keep count on column index
			col_num = 0
			# loop through found names
			for oc in names_xlsx:
				
				# keep count on column index
				col_num += 1

				if (oc == None):
					continue
				if (oc == 'Temperature (K)'):
					temper_indx = col_num
					self.TEMP = []
					self.tempt = []
					continue
				if (oc == 'RH (0-1)'):
					rh_indx = col_num
					self.RH = []
					self.RHt = []
					continue
				# if an observed particulate 
				# matter component
				if 'PM' in oc or 'pm' in oc:
			
					# assume that observations
					# account for any dilution,
					# so both particle components
					# (in ode_solv) and particle
					# number concentration (in
					# ode_updater)
					# are not subject to any
					# further dilution
					self.pp_dil = 0

					# if a mole fraction and 
					# first time a mole fraction 
					# seen
					if ('seedx' in oc and seedx_flag == 0):
						seedx_flag = 1
					if ('pconc' in oc and pconc_flag == 0):
						pconc_flag = 1
					if ('mean_rad' in oc):
						mean_rad_flag = 1
					if ('seedx' in oc):
						# index where size bin 
						# name starts
						strfi = oc.index('pm')+2
						# index where size bin name
						# finishes
						strei = -1*(oc[::-1].index('_')+1)
						
						# size bin name
						sb_nam.append(oc[strfi:strei])
						# hold component name
						pm_comp.append(oc[0:strfi-3])
						# hold column index
						pm_col.append(col_num)
						
					if ('pconc' in oc):
						# index where size bin 
						# name starts
						strfi = oc.index('pm')+2
						# size bin name
						sb_nam_pconc.append(oc[strfi::])
						# hold column index
						pconc_col.append(col_num)
					# mean radius of size bin (um)
					if ('mean_rad' in oc):
						# index where size bin 
						# name starts
						strfi = oc.index('pm')+2
						# size bin name
						sb_nam_mean_rad.append(oc[strfi::])
						# hold column index
						mean_rad_col.append(col_num)
					continue
	
				# if none of the above, then 
				# assume it's a gas-phase 
				# chemical component
				self.obs_comp.append(oc)
				comp_indx.append(col_num)
			
			# in case of simulation preparation mode, set
			# these components as ones to track change 
			# tendencies of
			if (self.sim_ci_file != []):
				self.dydt_trak = self.obs_comp
					
			# get number of components to consider
			nc_obs = len(self.obs_comp)
			# prepare to store observations, not 
			# forgetting a column for time
			self.obs = np.zeros((1, nc_obs+1))
			comp_indx = np.array((comp_indx)).astype('int')
			rh_indx = np.array((rh_indx)).astype('int')
			temper_indx = np.array((
				temper_indx)).astype('int')
		else:
			# if on first time
			if (ic == 1 and seedx_flag == 1):
				
				# prepare to hold seed component names
				self.seed_name = []
				# get unique names of seed components
				for seedi in pm_comp:
					if seedi not in self.seed_name:
						self.seed_name.append(seedi)
						
				# get number of unique components
				# with PM mole fractions
				uni_comp = len(self.seed_name)
				# get number of unique size bins
				uni_sb = len(np.unique(sb_nam))
				self.seedx = np.zeros((uni_comp, uni_sb, 1))
				# rearrange pm_col so that components
				# in columns and size bins in rows
				pm_col = np.array((pm_col)).reshape(uni_sb, uni_comp)

			if (ic == 1 and pconc_flag == 1):
				# get number of unique size bins
				uni_sb_pconc = len(np.unique(sb_nam_pconc))
				self.pconc = np.zeros((uni_sb_pconc, 1))
				# times of particle number concentration (s)
				self.pconct = np.zeros((1, 1))
				

			if (ic == 1 and mean_rad_flag == 1):
				# get number of unique size bins
				uni_sb_mean_rad = len(np.unique(sb_nam_mean_rad))
				self.mean_rad = np.zeros((uni_sb_mean_rad, 1))
				

			if (ic > 1): # if past the first time
				# add row onto data array
				self.obs = np.concatenate((self.obs, 
					(np.zeros((1, nc_obs+1)))), 
					axis=0)
				# add onto mole fraction array
				self.seedx = np.concatenate((self.seedx, 
					(np.zeros((uni_comp, 
					 uni_sb, 1)))), 
					axis=2)
				# add onto PM number concentration array
				self.pconc = np.concatenate((self.pconc, 
					(np.zeros((uni_sb_pconc, 1)))), 
					axis=1)
				self.pconct = np.concatenate((self.pconct, 
					(np.zeros((1, 1)))), 
					axis=1)
				# add onto mean radius array
				self.mean_rad = np.concatenate((self.mean_rad, 
					(np.zeros((uni_sb_mean_rad, 1)))), 
					axis=1)
			
			self.obs[ic-1, 1::] = np.array((i))[
				0:col_num+1][comp_indx]
			# get times (s)
			self.obs[ic-1, 0] = np.array((i))[0]
			
			# temperature and relative humidity observations
			if (temper_indx != -2):
				self.TEMP.append(
				np.array((i))[temper_indx])
				self.tempt.append(i[0])
			if (rh_indx != -2):
				self.RH.append(np.array((i))[rh_indx])
				self.RHt.append(i[0])

			# particulate matter observations at this time
			if (seedx_flag == 1):
				# loop through size bins
				for sbi in range(uni_sb):
					self.seedx[:, sbi, -1] = np.array((i))[
					0:col_num+1][pm_col[sbi, :]]
			# PM number concentration at this time (# particles/cm3)
			if (pconc_flag == 1):
				self.pconc[:, -1] = np.array((i))[0:col_num+1][pconc_col]
				self.pconct[0, -1] = np.array((i))[0]
			
			# mean radius of size bins at this time (um)
			if (mean_rad_flag == 1):
				self.mean_rad[:, -1] = 	np.array((i))[0:col_num+1][mean_rad_col]		

		ic += 1 # count on row iteration
	

	# whether particle number concentrations expressed 
	# by modes (0) or explicitly per size bin (1)
	if (self.pconc.shape[0] == self.num_asb):
		self.pmode = 1
	else:
		self.pmode = 0
	
	# ensure array defining whether particle injection
	# is continuous or instantaneous is same length as
	# number of injection times
	self.pcont = (np.zeros((1, self.pconct.shape[1]))).astype('int')

	# ensure that seed component names are
	# consistent with those in chemical scheme
	try:
		self.seed_name[self.seed_name.index('an')] = 'AMM_NIT'
		self.seed_name[self.seed_name.index('as')] = 'AMM_SUL'
		self.seed_name[self.seed_name.index('po')] = 'pri_org'
		self.seed_name[self.seed_name.index('ulvoc')] = 'sec_org-2'
		self.seed_name[self.seed_name.index('elvoc')] = 'sec_org-1'
		self.seed_name[self.seed_name.index('lvoc')] = 'sec_org0'
		self.seed_name[self.seed_name.index('svoc')] = 'sec_org1'
		self.seed_name[self.seed_name.index('ec')] = 'bc'
	except:
		self.seed_name = self.seed_name
	
	# ensure numpy array for rh arrays
	self.RH = np.array((self.RH)).astype('float')
	self.RHt = np.array((self.RHt)).astype('float')
	wb.close() # close excel file
	
	
	if hasattr(self, 'sim_ci_file'):
		self.ci_array = np.zeros((len(self.obs_comp)+1, 
			1)).astype('str')
		self.ci_array[0, 0] = 'molec/cm3/s' # units
		self.ci_array[1::, 0] = self.obs_comp # component names
	
	# get indices of components with concentrations to fit 
	# observations
	self.obs_comp_i = np.zeros((len(self.obs_comp))).astype('int')

	return(self)
