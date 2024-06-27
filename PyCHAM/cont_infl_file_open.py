########################################################################
#								       #
# Copyright (C) 2018-2024					       #
# Simon O'Meara : simon.omeara@manchester.ac.uk			       #
#								       #
# All Rights Reserved.                                                 #
# This file is part of PyCHAM                                          #
#                                                                      #
# PyCHAM is free software: you can redistribute it and/or modify it    #
# under the terms of the GNU General Public License as published by    #
# the Free Software Foundation, either version 3 of the License, or    #
# (at  your option) any later version.                                 #
#                                                                      #
# PyCHAM is distributed in the hope that it will be useful, but        #
# WITHOUT ANY WARRANTY; without even the implied warranty of           #
# MERCHANTABILITY or## FITNESS FOR A PARTICULAR PURPOSE.  See the GNU  #
# General Public License for more details.                             #
#                                                                      #
# You should have received a copy of the GNU General Public License    #
# along with PyCHAM.  If not, see <http://www.gnu.org/licenses/>.      #
#                                                                      #
########################################################################
'''opening and reading any user-provided continuous influx file'''
# called from mod_var_read.py

# define function to read in values relevant to continuous influxes
def cont_infl_open(self):

	# inputs: ---------------------------
	# self - the PyCHAM object of objects
	# -----------------------------------
	
	import openpyxl
	import os
	import numpy as np

	# size bin names
	sb_nam = []
	# names of particulate matter component
	pm_comp = []
	# size bin names for particle components
	sb_nam = []
	# size bin names for pconc
	sb_nam_pconc = []
	# size bin names for mean_rad
	sb_nam_mean_rad = []
	# prepare to hold seed component names
	self.seed_name = []
	
	if ('xls' in self.const_infl_path):
		
		try: # try to open the file at the user-supplied path

			try:
				wb = openpyxl.load_workbook(filename = 
					self.const_infl_path)
				
			except:
				const_infl_path0 = self.const_infl_path
				# see if present inside same folder as 
				# model variables
				# strip path to model variables file 
				# back to home directory
				try:
					path_start = -1*self.inname[::-1].index('/')
				except:
					path_start = -1*self.inname[::-1].index('\\')
				# path to file
				self.const_infl_path = str(self.inname[0:path_start] + 
					self.const_infl_path)
				
				wb = openpyxl.load_workbook(filename = self.const_infl_path)				
			
			
			try:
				sheet = wb['cont_infl']
			except:
				sheet = wb['const_infl']
			
			# component names are in first column, 
			# continuous influxes are in 
			# following columns		
			ir = -1 # count on row iteration
	
			# prepare to store component names and 
			# continuous influxes
			value = ''

			# names of gas-phase components with
			# continuous influx
			self.con_infl_nam = np.empty((0))	
			
			# loop through rows
			for i in sheet.iter_rows(values_only=True):
				
				ir += 1 # count on row iteration
				# header provides unit of emission rate 
				# and times
				
				if (ir == 0):
					self.abun_unit = str(i[0])	
			
					if ('ppb' not in 
					self.abun_unit):
						if ('mol' not in 
						self.abun_unit and 
						'cm' not in 
						self.abun_unit):
							self.err_mess = str('Error: units of continuous influx in first column of first row of the file for continuous influx of components could not be found, acceptable units are ppb or molec/cm3/s; file path attempted was: ' + self.const_infl_path)
							
							return(self)

					clim = 0 # count on columns
					for ic in i[0::]:	
						if ic is None:
							# stop looping 														# through 
							# columns
							break
						# count on columns	
						clim +=1 
					
					# if looping over a 24 hour 
					# period
					# then limit influxes to the
					# first 
					# provided 24 hours
					if (self.con_infl_tf == 1): 
						clim = sum(i[1:clim] < 24.*3.6e3)
					
					self.con_infl_t = np.array((i[1:clim])).astype('float')
					self.con_infl_C = np.empty((0, clim-1)).astype('float')
					# number of times
					nt = len(self.con_infl_t)
					# remember times for particle influx
					self.pconct = np.array((self.con_infl_t)).reshape(1, -1)
	
					# prepare to hold seed mole fractions per 
					# component (rows), per size bin 
					# (columns), per time (3rd dimension)
					self.seedx = np.zeros((0, 0, nt))

					# initiate array for holding particle number
					# concentrations (particles/cm3)
					self.pconc = np.zeros((0, nt))

					# initiate array for holding particle mean
					# radius (um)
					self.mean_rad = np.zeros((0, nt))

					continue # onto next row			

				# get names of components (matching chemical scheme names) 
				# and their continuous influx rates (abundance unit given above)
				else:
					# reached end of contiguous components
					if (i[1] is None):
						break
			
					# if a particulate 
					# matter component
					if 'PM' in i[0] or 'pm' in i[0] or 'sb' in i[0]:
						oc = i[0]
						if ('seedx' in oc):
							# index where size bin 
							# name starts
							strfi = oc.index('pm')+2
							# index where size bin name
							# finishes
							strei = -1*(oc[::-1].index('_')+1)
						
							# size bin name
							sb_now = oc[strfi:strei]

							# hold unique size bin names
							if sb_now not in sb_nam:
								sb_nam.append(sb_now)
								# extend matrix for values
								self.seedx = np.concatenate((
								self.seedx, np.zeros((
								self.seedx.shape[0], 1, nt))), 
								axis=1)

							# size bin index
							sbi = sb_nam.index(sb_now)

							# hold component name
							pm_comp = (oc[0:strfi-3])
					
							# hold unique names of seed components
							if pm_comp not in self.seed_name:
								self.seed_name.append(pm_comp)
								# extend matrix for values
								self.seedx = np.concatenate((
								self.seedx, np.zeros((
								1, 1, nt))), 
								axis=0)
							
							# component index
							cii = self.seed_name.index(pm_comp)

							# store the mole fractions of
							# this component in this size bin
							self.seedx[cii, sbi, :] = i[1:clim]

						if ('pconc' in oc):
							self.pconc = np.concatenate((
							self.pconc, np.zeros((1, nt))), axis=0)
							self.pconc[-1, :] = i[1:clim]

						# mean radius of size bin (um)
						if ('mean_rad' in oc):
							self.mean_rad = np.concatenate((
							self.mean_rad, np.zeros((1, nt))), 
							axis=0)
							self.mean_rad[-1, :] = i[1:clim]

						continue

					# if a gas-phase component name
					self.con_infl_nam = np.concatenate((
					self.con_infl_nam, np.array((str(i[0]))).reshape(1)))

					# continuous influx rate
					self.con_infl_C = np.concatenate((
					self.con_infl_C, np.array((i[1:clim])).reshape(
					1, -1).astype('float')))
	
			wb.close() # close excel file
			
		except: # if file not found tell user
			self.err_mess = str('Error: file provided by user in model variables file for continuous influx of components was either not found or could not be opened, file path attempted was: ' + const_infl_path0)

	else:
		self.con_infl_nam = 'not in a file'

	# whether particle number concentrations expressed 
	# by modes (0) or explicitly per size bin (1)
	if (self.pconc.shape[0] == self.num_asb):
		self.pmode = 1
	else:
		self.pmode = 0

	# ensure array defining whether particle injection
	# is continuous or instantaneous is same length as
	# number of injection times
	self.pcont = (np.ones((1, self.pconct.shape[1]))).astype('int')
	
	return(self)
