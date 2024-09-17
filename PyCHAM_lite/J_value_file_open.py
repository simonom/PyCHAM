########################################################################
#								       #
# Copyright (C) 2018-2024					       #
# Simon O'Meara : simon.omeara@manchester.ac.uk			       #
#								       #
# All Rights Reserved.                                                 #
# This file is part of PyCHAM                                          #
#                                                                      #
# PyCHAM is free software: you can redistribute it and/or modify it    #
# under the terms of the GNU General Public License as published by    #
# the Free Software Foundation, either version 3 of the License, or    #
# (at  your option) any later version.                                 #
#                                                                      #
# PyCHAM is distributed in the hope that it will be useful, but        #
# WITHOUT ANY WARRANTY; without even the implied warranty of           #
# MERCHANTABILITY or## FITNESS FOR A PARTICULAR PURPOSE.  See the GNU  #
# General Public License for more details.                             #
#                                                                      #
# You should have received a copy of the GNU General Public License    #
# along with PyCHAM.  If not, see <http://www.gnu.org/licenses/>.      #
#                                                                      #
########################################################################
'''opening and reading any user-provided photolysis rate files'''
# called from mod_var_read.py

# define function
def J_value_file_open(path, self):

	# inputs: --
	# self - PyCHAM object
	# ----------

	import openpyxl
	import os
	import numpy as np

	try: # in case full path provided
		
		wb = openpyxl.load_workbook(filename = path)
	except: # in case in same folder as model variables file
		# note if this fails, then the except condition
		# under the light_staus intepretation part of
		# mod_var_read is invoked
		pd_indx = self.inname[::-1].index('/')
		pd = self.inname[0:-pd_indx]
		path = str(pd + path)
		wb = openpyxl.load_workbook(filename = path)
	
	sheet = wb['J']
	# time (seconds) is in first row, J values
	# are in later rows
	ir = -1 # count on row iteration
	
	# loop through rows
	for i in sheet.iter_rows(values_only = True):
		
		# count on rows
		ir += 1

		if (ir == 0):
			# get times for lighting (s through expeirment)
			self.light_time = np.array((i[0::]))
			# prepare for variable names
			self.stored_J = np.zeros((0, len(self.light_time)))
			
		else: # if onto J valuess

			self.stored_J = np.concatenate((
			self.stored_J, np.zeros((1, 
			len(self.light_time)))), axis=0)

			self.stored_J[-1, :] = i[0::]

	return(self)
