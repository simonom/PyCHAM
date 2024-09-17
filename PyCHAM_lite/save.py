########################################################################
#								       #
# Copyright (C) 2018-2024					       #
# Simon O'Meara : simon.omeara@manchester.ac.uk			       #
#								       #
# All Rights Reserved.                                                 #
# This file is part of PyCHAM                                          #
#                                                                      #
# PyCHAM is free software: you can redistribute it and/or modify it    #
# under the terms of the GNU General Public License as published by    #
# the Free Software Foundation, either version 3 of the License, or    #
# (at  your option) any later version.                                 #
#                                                                      #
# PyCHAM is distributed in the hope that it will be useful, but        #
# WITHOUT ANY WARRANTY; without even the implied warranty of           #
# MERCHANTABILITY or## FITNESS FOR A PARTICULAR PURPOSE.  See the GNU  #
# General Public License for more details.                             #
#                                                                      #
# You should have received a copy of the GNU General Public License    #
# along with PyCHAM.  If not, see <http://www.gnu.org/licenses/>.      #
#                                                                      #
########################################################################
'''module to save PyCHAM results to files'''
# module called following simulation in PyCHAM to store generated and 
# input information

import os
import sys
import numpy as np
import csv
from shutil import copyfile
import pickle

def saving(y_mat, Nresult_dry, Nresult_wet, t_out, num_comp, 
	Cfactor_vst, testf, numsb, y_mw, MV,
	time_taken, x2, rbou_rec, rbou00, upper_bin_rad_amp, 
	indx_plot, H2Oi, siz_str, cham_env, self):
	
	# inputs: ----------------------------------------------------------------------------
	
	# y_mat - component (columns) concentrations with time (rows) (# molecules/cm3 (air))
	# Nresult_dry  - number concentration of dry particles 
	# per size bin (# particles/cm3 (air))
	# Nresult_wet  - number concentration of dry particles 
	# per size bin (# particles/cm3 (air))
	# Cfactor - conversion factor to change gas-phase concentrations from # molecules/cm3 
	# (air) into ppb
	# testf - flag to show whether in normal mode (0) or test mode (1)
	# numsb - number of size bins
	# self.sav_nam - folder name/path to save results to
	# self.dydt_vst - tendency to change of user-specified 
	# components
	# self.dydt_trak - user-input names of components to track
	# self.comp_namelist - names of components given by the 
	# chemical scheme file
	# upper_bin_rad_amp - factor upper bin radius found increased 
	#	by in Size_distributions.py for more than 1 size bin, 
	# 	or in pp_intro.py for 1 size bin
	# Cfactor_vst - one billionth the molecular concentration in a 
	#	unit volume of chamber (# molecules/cm3) per recording 
	# time step
	# time_taken - computer processing time for entire simulation 
	#	(s)
	# y_mw - molecular weights (g/mol)
	# self.nom_mass - nominal molar mass (g/mol)
	# MV - molar volumes of all components (cm3/mol)
	# time_taken - simulation computer time (s)
	# self.seed_name - chemical scheme name of component comprising seed particles
	# x2 - record of size bin radii (um)
	# rbou_rec - radius bounds per size bin (columns) per time step (rows) (um)
	# self.wall_on - marker for whether wall on or off
	# self.space_mode - type of spacing between particle size bins (log or lin)
	# rbou00 - initial lower size (radius) bound of particles (um)
	# upper_bin_rad_amp - factor increase in radius of upper bound
	# indx_plot - index of components to plot gas-phase 
	# concentration temporal profiles of in 
	# standard results plot
	# self.comp0 - names of components to plot gas-phase 
	# 	concentration temporal profiles of in standard results plot
	# self.yrec_p2w - concentration of components on the wall due to 
	#	particle-wall loss, stacked by component first then by
	#	size bin (# molecules/cm3)
	# self.sch_name - path to chemical scheme file
	# self.inname - path to model variables file
	# self.rel_SMILES - SMILES strings for components in chemical scheme
	# self.Psat_Pa_rec - pure component saturation vapour pressures at 298.15 K (Pa)
	# self.Psat_rec0 - pure component saturation vapour pressures 
	#	at starting temperature of simulation (# molecules/cm3)
	# self.OC - oxygen to carbon ratio of components
	# self.HC - hydrogen to carbon ratio of components
	# H2Oi - index of water
	# self.seedi - index of seed components
	# siz_str - the size structure
	# cham_env - chamber environmental conditions (temperature (K), 
	# pressure (Pa) and relative humidity
	# self.RO2_indices[:, 1] - alkyl peroxy radical indices
	# self.RO_indx - alkoxy radical indices
	# self.tot_in_res_ft - record of cumulative inputs of injected components (ug/m3)
	# self.tf - transmission factor of light, possibly segregated by wavelength
	# self - reference to PyCHAM
	# ---------------------------------------------------------------
	
	# correct for changes to size bin radius bounds
	if ((numsb-self.wall_on) > 0):
		rbou_rec[:, 0] = rbou00
		rbou_rec[:, -1] = rbou_rec[:, -1]/upper_bin_rad_amp

	if (testf == 1):
		return(0) # return dummy

	# note that the directory to save results to (self.output_by_sim) 
	# is created in user_input.py

	# create folder to store copies of inputs
	os.makedirs(str(self.output_by_sim+'/inputs'))
	# making a copy of the chemical scheme and model variables input files
	if '/' in self.sch_name: # if using / to separate path parts
		output_by_sim_sch_ext = str(self.output_by_sim + '/inputs/' + self.sch_name.split('/')[-1])
	if '\\' in self.sch_name: # if using \\ to separate path parts
		output_by_sim_sch_ext = str(self.output_by_sim + '/inputs/' + self.sch_name.split('\\')[-1])
	
	output_by_sim_mv_ext = str(self.output_by_sim + '/inputs/' + self.inname.split('/')[-1])
	# if not in default model variables mode or in automatic calling mode
	if (self.inname != 'Default' and type(self.param_const) != dict):
		copyfile(self.inname, output_by_sim_mv_ext)
	
	# if in automatic calling mode, then save the specified model variables
	if (type(self.param_const) == dict):
		output_by_sim_mv_ext = str(self.output_by_sim + '/inputs/specified_model_variables.txt')
		with open(output_by_sim_mv_ext, 'w') as f:
			for key, value in self.param_const.items():
				f.write('%s = %s\n' % (key, value))
		f.close()
	
	# saving dictionary
	# dictionary containing variables for model and components
	const = {}
	const["number_of_size_bins"] = numsb
	const["number_of_components"] = num_comp
	#const["molecular_weights_g/mol_corresponding_to_component_names"] = (np.squeeze(y_mw[:, 0]).tolist())
	#const["nominal_molar_mass_g/mol"] = self.nom_mass.tolist()
	#const["molar_volumes_cm3/mol"] = (MV[:, 0].tolist())
	#const["organic_peroxy_radical_index"] = (self.RO2_indices[:, 1].tolist())
	#const["organic_alkoxy_radical_index"] = self.RO_indx
	#const["organic_HOM_peroxy_radical_index"] = self.HOMRO2_indx
	#const["organic_HOMs_index"] = self.HOMs_indx	
	#const["chem_scheme_names"] = self.comp_namelist
	#const["SMILES"] = self.rel_SMILES
	const["factor_for_multiplying_ppb_to_get_molec/cm3_with_time"] = (Cfactor_vst.tolist())
	const["simulation_computer_time(s)"] = time_taken
	const["seed_name"] = self.seed_name
	const["wall_on_flag_0forNO_>0forYES"] = self.wall_on
	const["space_mode"] = self.space_mode
	#const["pure_component_saturation_vapour_pressures_at_298.15K_Pa"] = self.Psat_Pa_rec.tolist()
	#const["oxygen_to_carbon_ratios_of_components"] = self.OC.tolist()
	#const["hydrogen_to_carbon_ratios_of_components"] = self.HC.tolist()
	const["index_of_water"] = H2Oi
	const["index_of_seed_components"] = self.seedi.tolist()
	const["size_structure_0_for_moving_centre_1_for_full_moving"] = siz_str
	const["output_by_sim_sch_ext"] = output_by_sim_sch_ext
	const["output_by_sim_mv_ext"] = output_by_sim_mv_ext
	
	with open(os.path.join(self.output_by_sim,'model_and_component_constants'),'w') as f:
		for key in const.keys():
			f.write("%s,%s\n"%(key, const[key]))
	
	# save
	save_path = str(self.output_by_sim + '/nom_mass') # path
	np.save(save_path, self.nom_mass, allow_pickle=True)
	
	save_path = str(self.output_by_sim + '/y_mw') # path
	np.save(save_path, np.squeeze(y_mw), allow_pickle=True)

	save_path = str(self.output_by_sim + '/MV') # path
	np.save(save_path, np.squeeze(MV), allow_pickle=True)

	save_path = str(self.output_by_sim + '/comp_namelist') # path
	np.save(save_path, self.comp_namelist, allow_pickle=True)

	save_path = str(self.output_by_sim + '/rel_SMILES') # path
	np.save(save_path, self.rel_SMILES, allow_pickle=True)

	save_path = str(self.output_by_sim + '/pure_component_saturation_vapour_pressures_at_298p15K_Pa') # path
	np.save(save_path, self.Psat_Pa_rec, allow_pickle=True)

	save_path = str(self.output_by_sim + '/pure_component_saturation_vp_at_startT_molec_percm3') # path
	np.save(save_path, self.Psat_rec0, allow_pickle=True)

	save_path = str(self.output_by_sim + '/oxygen_to_carbon_ratios_of_components') # path
	np.save(save_path, self.OC, allow_pickle=True)
	
	save_path = str(self.output_by_sim + '/hydrogen_to_carbon_ratios_of_components') # path
	np.save(save_path, self.HC, allow_pickle=True)

	save_path = str(self.output_by_sim + '/OOH_index') # path
	np.save(save_path, self.OOH, allow_pickle=True)
 
	save_path = str(self.output_by_sim + '/HOM_OOH_index') # path
	np.save(save_path, self.HOM_OOH, allow_pickle=True)

	save_path = str(self.output_by_sim + '/OH_index') # path
	np.save(save_path, self.OH, allow_pickle=True)

	save_path = str(self.output_by_sim + '/HOM_OH_index') # path
	np.save(save_path, self.HOM_OH, allow_pickle=True)

	save_path = str(self.output_by_sim + '/carbonyl_index') # path
	np.save(save_path, self.carbonyl, allow_pickle=True)

	save_path = str(self.output_by_sim + '/HOM_carbonyl_index') # path
	np.save(save_path, self.HOM_carbonyl, allow_pickle=True)

	save_path = str(self.output_by_sim + '/NO3_index') # path
	np.save(save_path, self.NO3, allow_pickle=True)

	save_path = str(self.output_by_sim + '/HOM_NO3_index') # path
	np.save(save_path, self.HOM_NO3, allow_pickle=True)

	save_path = str(self.output_by_sim + '/organic_peroxy_radical_index') # path
	np.save(save_path, self.RO2_indices[:, -1], allow_pickle=True)	
	
	save_path = str(self.output_by_sim + '/organic_alkoxy_radical_index') # path
	np.save(save_path, self.RO_indx, allow_pickle=True)

	save_path = str(self.output_by_sim + '/organic_HOM_peroxy_radical_index') # path
	np.save(save_path, self.HOM_RO2_indx, allow_pickle=True)

	save_path = str(self.output_by_sim + '/organic_HOMs_index') # path
	np.save(save_path, self.HOMs_indx, allow_pickle=True)
	
	save_path = str(self.output_by_sim + '/organic_ROOR_index') # path
	np.save(save_path, self.ROOR_indx, allow_pickle=True)

	# saving Ademipo's group indices for Baker et al. 2024 (
	# doi.org/10.5194/acp-24-4789-2024) product classes start --------
	save_path = str(self.output_by_sim + '/ROORBaker_indx') # path
	np.save(save_path, self.HOMFragBaker_indx, allow_pickle=True)

	save_path = str(self.output_by_sim + '/HOMFragBaker_indx') # path
	np.save(save_path, self.ROORBaker_indx, allow_pickle=True)

	save_path = str(self.output_by_sim + '/HOMRO2Baker_indx') # path
	np.save(save_path, self.HOMRO2Baker_indx, allow_pickle=True)

	save_path = str(self.output_by_sim + '/HOMMonBaker_indx') # path
	np.save(save_path, self.HOMMonBaker_indx, allow_pickle=True)

	# Baker et al. 2024 product classes end --------------------------

	# store concentrations of all components in all phases at the
	# final time, in case the next simulation wants to use these
	# (molecules/cm3 (air))
	self.yorig = y_mat[-1, :]

	# convert gas-phase concentrations from # molecules/cm3 (air) into ppb
	# leaving any particle or wall concentrations as # molecules/cm3 (air)
	y_mat[:, 0:num_comp] = y_mat[:, 0:num_comp]/(Cfactor_vst.reshape(len(Cfactor_vst), 1))
	
	y_header = str('') # prepare header for concentrations with time file 
	x2_header = str('') # prepare header for files relating to size bins
	
	for i in range(numsb+1): # loop through size bins

		if i == 0:
			end = '_g'
		if ((i > 0) and (i < numsb)):
			end = str('_p' + str(i))
			x2_header = str(x2_header + str(i))
		if (i == numsb):
			if (self.wall_on == 0):
				end = str('_p'  + str(i))
				x2_header = str(x2_header+str(np.repeat(i, num_comp)))
			else:
				end = '_w'
		for ii in range(num_comp):
			if i  == 0 and ii == 0:
				start = ''
			else:
				start = ', '
			y_header = str(y_header + str(start + self.comp_namelist[ii]) + end)
			
	# saving both gas, particle and wall concentrations of components
	np.savetxt(os.path.join(self.output_by_sim, 'concentrations_all_components_all_times_gas_particle_wall'), y_mat, delimiter=',', header=str('time changes with rows which correspond to the time output file, components in columns, with _g representing gas phase (ppb), _pi representing particle phase where i is the size bin number (starting at 1) (molecules/cm3 (air)) and _w is the wall phase (molecules/cm3 (air))\n' + y_header)) 		
	
	# saving time of outputs
	np.savetxt(os.path.join(self.output_by_sim, 'time'), t_out, delimiter=',', header='time (s), these correspond to the rows in the concentrations_all_components_all_times_gas_particle_wall, particle_number_concentration and size_bin_radius output files')
	
	# saving environmental conditions (temperature, pressure, relative humidity, transmission factor)
	# prepare to concatenate transmission factor of light (0-1)
	cham_env = cham_env.astype('str')

	np.savetxt(os.path.join(self.output_by_sim, 'chamber_environmental_conditions'), cham_env, fmt = '%s', delimiter=',', header='chamber environmental conditions throughout the simulation, with rows corresponding to the time points in the time output file, first column is temperature (K), second is pressure (Pa), third is relative humidity (fraction (0-1), fourth is transmission factor of light (0-1))')
	
	# saving the index and names of components whose gas-phase 
	# temporal profiles can be plotted on the standard results plot
	fname = os.path.join(self.output_by_sim, 'components_with_initial_gas_phase_concentrations_specified')
	
	np.savetxt(fname, [indx_plot, self.comp0], delimiter =', ', header='index (top row) and chemical scheme name (bottom row) of components with initial gas-phase concentrations specified', fmt ='% s') 
	
	# if tracking of tendencies to change requested by user, 
	# loop through the components
	# and save the tendency record for each of these (%/hour)
	if (len(self.dydt_vst) > 0):
		compind = 0
		# get names of tracked components
		dydtnames = self.dydt_vst['comp_names'] 
		
		if hasattr(self, 'ci_array'):
			if (len(self.ci_array) > 0):
				# times (s)
				self.ci_array = np.concatenate((
					self.ci_array, 
					np.zeros((self.
					ci_array.shape[0], 
					len(t_out)-1)).astype('str')), 
					axis=1)
				self.ci_array[0, 1::] = t_out[
					0:-1].astype(str)
			

		# loop through components to record the tendency of 
		# change 
		for comp_name in dydtnames:
			# open relevant dictionary value, to get the 
			# 2D numpy array for saving
			dydt_rec = np.array(self.dydt_vst[
				str(str(comp_name) + '_res')])
			
			# get user-input name of this component
			comp_name = str(self.dydt_trak[compind] +
				'_rate_of_change')
			# save
			np.savetxt(os.path.join(self.output_by_sim, comp_name), dydt_rec, delimiter=',', header='tendency to change, top row gives equation number (where number 0 is the first equation), 3rd column from end is gas-particle partitioning, 2nd column from end is gas-wall partitioning, final column is dilution (molecules/cm3/s (air))')

			if hasattr(self, 'ci_array'):
				if (len(self.ci_array) > 0):
					# get relevant row
					ri = self.ci_array[:, 0] == str(
						self.dydt_trak[compind])

					# influx rate
					self.ci_array[ri, 
					1::] = self.dydt_vst[
					str(str(
					self.dydt_trak[compind]) + 
					'_ci')][:, 0]
				
			compind += 1

		# in case writing continuous influx to a spreadsheet
		if (self.sim_ci_file != []):
			from openpyxl.utils.cell import\
			get_column_letter

			# writing to excel file
			from openpyxl import Workbook
			wb = Workbook()
			ws = wb.active
			ws.title = 'cont_infl'
			# loop through columns in first row
			for ic in range(self.ci_array.shape[1]):
				col_lett = get_column_letter(ic+1)
				# cell code
				cc = str(col_lett + '1')
				ws[cc] = self.ci_array[0, ic]
			# loop through other rows
			for ir in range(1, self.ci_array.shape[0]):
				ws.append(self.ci_array[ir, :].tolist())
			# parent directory of model variables 
			# file
			pd_indx = self.inname[::-1].index('/')
			pd = self.inname[0:-pd_indx]
			# save
			wb.save(pd + self.sim_ci_file)
		
	# saving generation of components
	np.savetxt(os.path.join(self.output_by_sim, 'component_generation'), self.gen_num, delimiter=',', header='generation number of each component (where the initial unoxidised Vself.OC is generation number 0), with the order corresponding to that of components in the concentrations_all_components_all_times_gas_particle_wall file.')
	
	if ((numsb-self.wall_on) > 0): # if particles present
	
		# saving the concentration of components on the wall due to particle deposition to wall
		np.savetxt(os.path.join(self.output_by_sim, 'concentrations_all_components_all_times_on_wall_due_to_particle_deposition_to_wall'), self.yrec_p2w, delimiter=',', header=str('concentration of components on wall due to particle deposition to wall (# molecules/cm3 (air)) time changes with rows which correspond to the time output file, components in columns and size bin changing with columns with size bin numbers given in the second row of the header\n'+ x2_header)) 
	
		np.savetxt(os.path.join(self.output_by_sim, 'particle_number_concentration_dry'), Nresult_dry, delimiter=',',
				header=('particle number concentration assuming water removed from particles (#/cm3 (air)), with time changing with rows (corresponding times given in the time output file) and size bin changing with columns with size bin numbers given in the second row of the header\n'+x2_header))
		
		np.savetxt(os.path.join(self.output_by_sim, 'particle_number_concentration_wet'), Nresult_wet, delimiter=',',
				header=('particle number concentration assuming water not removed from particles (#/cc (air)), with time changing with rows (corresponding times given in the time output file) and size bin changing with columns with size bin numbers given in the second row of the header\n'+x2_header))	
	
		np.savetxt(os.path.join(self.output_by_sim, 'size_bin_radius'), x2, delimiter=',',
				header= str('particle radii (um) per size_bin (including water contribution to size), with size bins represented by columns and their number (starting from 1) given in second line of header, per time step which is represented by rows and corresponding times given in the time output file \n'+x2_header))
	
		np.savetxt(os.path.join(self.output_by_sim, 'size_bin_bounds'), rbou_rec, delimiter=',',
				header=str('particle size bin bounds (um), with size bins represented by columns and their number (starting at 1 and in line with the lower bound) given in second line of header, per time step which is is represented by rows and corresponding times given in the time output file \n'+x2_header))		
	
	np.savetxt(os.path.join(self.output_by_sim, 'total_concentration_of_injected_components'), self.tot_in_res_ft, delimiter=',',
				header=str('the total concentration (ug/m3) of injected (through initial gas-phase concentration, instantaneous and/or continuous gas-phase influx) components, with component index (relative to all components) in the first row and its cumulative injected concentrations in following rows'))
	
	# this save added on 05/05/2022
	# save self, so that same simulation can be recreated again
	import pickle
	# transfer variables in self object into variables in a list
	model_var_dict = {} # empty dictionary to hold wanted variables
	for attr, value in self.__dict__.items(): # loop through self object attributes
		# don't need to save QGridLayout objects
		if attr == 'title' or attr == 'left' or attr == 'top' or attr == 'width' or attr == 'height' or attr == 'initUI' or attr == 'setWindowTitle' or attr == 'setGeometry' or attr == 'setLayout':
			continue
		if '_orig' in attr or 'sch_name' in attr or 'xml_name' in attr or 'photo_path' in attr or 'chem_sch_mrk' in attr or 'TEMP' in attr or 'tempt' in attr or 'wall_on' in attr or 'update_stp' in attr or 'tot_time' in attr or 'save_step' in attr:
			# transfer to dictionary
			model_var_dict.update({attr:value})
	with open(os.path.join(self.output_by_sim, 'simulation_self.pickle'), "wb") as f:
		pickle.dump(model_var_dict, f)
	
	# if save name is the default, then remove to ensure no duplication in future
	if (self.sav_nam == 'default_res_name'):
		import shutil	
		shutil.rmtree(self.output_by_sim)
		
	# store group indices as in retr_out
	# empty dictionary to contain indices of certain groups of components
	group_indx = {}

	group_indx['OOH'] = self.OOH
	group_indx['HOM_OOH'] = self.HOM_OOH
	group_indx['OH'] = self.OH
	group_indx['HOM_OH'] = self.HOM_OH
	group_indx['carbonyl'] = self.carbonyl
	group_indx['HOM_carbonyl'] = self.HOM_carbonyl
	group_indx['NO3'] = self.NO3
	group_indx['HOM_NO3'] = self.HOM_NO3
	group_indx['RO2i'] = self.RO2_indices[:, -1]
	group_indx['ROi'] = self.RO_indx
	group_indx['HOMRO2'] = self.HOM_RO2_indx 
	group_indx['HOMs'] = self.HOMs_indx	
	group_indx['ROOR'] = self.ROOR_indx	

	if ((numsb-self.wall_on) == 0): # if zero size bins
		# ensure numpy arrays, rather than float
		Nresult_wet = (np.array((Nresult_wet))).reshape(-1, 1)
		Nresult_dry = (np.array((Nresult_dry))).reshape(-1, 1)
		rbou_rec = (np.array((rbou_rec))).reshape(-1, 1)
		
	if ((numsb-self.wall_on) == 1): # if one size bins
		# ensure numpy arrays, rather than float
		Nresult_wet = (np.array((Nresult_wet)))
		Nresult_dry = (np.array((Nresult_dry)))
		rbou_rec = (np.array((rbou_rec)))	
	
	# to ensure quick use of output, store results in an object in an identical way to retr_out
	# create a class to hold outputs
	class ro_outputs:
		sp = output_by_sim_sch_ext # chemical scheme path
		vp = output_by_sim_mv_ext # model variables path
		gi = group_indx # indices of groups of components
		gen_numbers = self.gen_num # for each component, the generation number
		HyC = self.HC.tolist() # hydrogen:carbon ratios for each component, this output added on 31/05/2022
		nominal_mass = self.nom_mass.tolist()
		nsb = numsb
		nc = num_comp
		cfac = Cfactor_vst
		yrec = y_mat
		Nrec_dry = Nresult_dry
		rad = rbou_rec
		cen_size = x2
		thr = t_out/3600.
		rSMILES = self.rel_SMILES
		comp_MW = y_mw
		Nrec_wet = Nresult_wet
		names_of_comp = self.comp_namelist
		comp_MV = MV
		proc_time = time_taken
		wf = self.wall_on
		spacing = self.space_mode
		plot_indx = indx_plot
		init_comp = self.comp0
		part_to_wall = self.yrec_p2w
		vpPa = self.Psat_Pa_rec
		vpPa0 = self.Psat_rec0
		O_to_C = self.OC.tolist()
		H2O_ind = H2Oi
		seed_ind = self.seedi.tolist()
		siz_struc = siz_str
		env_cond = cham_env
		total_influx = self.tot_in_res_ft
		
	self.ro_obj = ro_outputs() # create object to hold outputs
	
	return()
