##########################################################################################
#                                                                                        											 #
#    Copyright (C) 2018-2023 Simon O'Meara : simon.omeara@manchester.ac.uk                  				 #
#                                                                                       											 #
#    All Rights Reserved.                                                                									 #
#    This file is part of PyCHAM                                                         									 #
#                                                                                        											 #
#    PyCHAM is free software: you can redistribute it and/or modify it under              						 #
#    the terms of the GNU General Public License as published by the Free Software       					 #
#    Foundation, either version 3 of the License, or (at your option) any later          						 #
#    version.                                                                            										 #
#                                                                                        											 #
#    PyCHAM is distributed in the hope that it will be useful, but WITHOUT                						 #
#    ANY WARRANTY; without even the implied warranty of MERCHANTABILITY or FITNESS       			 #
#    FOR A PARTICULAR PURPOSE.  See the GNU General Public License for more              				 #
#    details.                                                                            										 #
#                                                                                        											 #
#    You should have received a copy of the GNU General Public License along with        					 #
#    PyCHAM.  If not, see <http://www.gnu.org/licenses/>.                                 							 #
#                                                                                        											 #
##########################################################################################
'''example code for interpolating actinic fluxes at known wavelengths into unit resolution and extrapolating to a defined range'''

# define function
def ex_interp():

	# import dependencies
	import openpyxl
	import numpy as np
 
	# location of the file containing original data
	path = "C:\\Users\\Psymo\\Desktop\\test.xlsx"
 
	# open the workbook
	# workbook object is created
	wb_obj = openpyxl.load_workbook(path)

	# make object out of first sheet
	sheet_obj = wb_obj.active

	# prepare to read in wavelengths (nm)
	wv = []

	# prepare to read in actinic fluxes (photons/cm2/nm/s)
	af = []

	# number of continuous rows from first containing values
	nb_row = sheet_obj.max_row

	# loop through rows
	for row in sheet_obj.iter_rows(min_row = 1, max_row = nb_row, max_col=2):
		col_cnt = 0 # count through columns
		for cell in row:
			if col_cnt == 0: # count through columns
				wv.append(cell.value) # wavelength
			if col_cnt == 1: # count through columns
				af.append(cell.value) # actinic flux
			col_cnt += 1 # count through columns

	# now create array containing unit resolution wavelength
	wv_new = (np.arange(254, 850, 1)).reshape(-1, 1)

	# create empty array to hold estimated actinic fluxes
	af_new = np.zeros((len(wv_new), 1))

	# prepare for short wavelengths
	if min(wv_new) < min(wv):
		# rate of change between two shortest 
		# provided wavelengths (actinic flux/wavelength)
		dfdw_sh = (af[1]-af[0])/(wv[1]-wv[0])

	# prepare for long wavelengths
	if max(wv) < max(wv_new):
		# rate of change between two longest 
		# provided wavelengths (actinic flux/wavelength)
		dfdw_lo = (af[-1]-af[-2])/(wv[-1]-wv[-2])
	

	# loop through new wavelengths to get estimated actinic fluxes
	for wvi in range(len(wv_new)):	
		
		if wv_new[wvi] < wv[0]: # if extrapolation needed
			af_new[wvi] = af[0]-dfdw_sh*(wv[0]-wv_new[wvi])
			
			if af_new[wvi] < 0: # ensure no negatives
				af_new[wvi] = 0.
			continue

		if wv_new[wvi] > wv[-1]: # if extrapolation needed
			af_new[wvi] = af[-1]+dfdw_lo*(wv_new[wvi]-wv[-1])	

			if af_new[wvi] < 0: # ensure no negatives
				af_new[wvi] = 0.

			continue

		# if interpolation needed

		# wavelength now
		wv_now = wv_new[wvi]
	
		# lower index to use
		li = sum((wv_now-wv)>0)-1
		
		# upper index to use
		ui = li+1	

		# get rate of change of actinic flux with wavelength
		dfdw = (af[ui]-af[li])/(wv[ui]-wv[li])
		af_new[wvi] = af[li]+dfdw*(wv_new[wvi]-wv[li])
	
		if af_new[wvi] < 0: # ensure no negatives
				af_new[wvi] = 0.
	
	# horizontally concatenate wavelengths and actinic fluxes
	res = np.concatenate((wv_new, af_new), axis=1)

	# save new results to csv file
	np.savetxt("C:\\Users\\Psymo\\Desktop\\PyCHAM\\PyCHAM\\PyCHAM\\photofiles\\new_af.csv", res, delimiter=",")
	# plot old and new values for a visual check
	import matplotlib.pyplot as plt
	plt.semilogy(wv, af, '-x', label = 'original')
	plt.semilogy(wv_new, af_new, '--', label = 'new')
	plt.legend()
	plt.show()
	
# call function
ex_interp()